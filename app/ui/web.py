
import logging
import os as _os
import sys
import builtins

# ── Logging 設定：所有 print 同步寫入 log 檔 ──
# web.py is at app/ui/web.py
# go up 3 levels to reach project root: ui -> app -> project root
CODE_DIR = _os.path.dirname(_os.path.abspath(__file__))
_APP_DIR = _os.path.dirname(CODE_DIR)
BASE_DIR = _os.path.dirname(_APP_DIR)

# 確保專案根目錄在 sys.path，讓 `from app.xxx import` 可以找到
if BASE_DIR not in sys.path:
    sys.path.insert(0, BASE_DIR)
_log_dir = _os.path.join(BASE_DIR, "logs")
_os.makedirs(_log_dir, exist_ok=True)

logging.basicConfig(
    filename=_os.path.join(_log_dir, "terminal.log"),
    level=logging.DEBUG,
    format="%(asctime)s [%(levelname)s] %(message)s",
    encoding="utf-8",
    filemode="a"   # a = append，不覆蓋
)

if not hasattr(builtins, "_original_print"):
    import functools
    setattr(builtins, "_original_print", getattr(builtins, "print"))
    
    @functools.wraps(getattr(builtins, "_original_print"))
    def custom_print(*args, **kwargs):
        getattr(builtins, "_original_print")(*args, **kwargs)
        logging.info(" ".join(str(a) for a in args))
    setattr(builtins, "print", custom_print)
# ─────────────────────────────────────────────

# ── 結構化 User Logger：記錄 user@IP，便於多人問題追蹤 ──
from app.utils.app_logger import get_user_logger

def _get_client_ip() -> str:
    """取得目前 Streamlit session 的 client IP（找不到則回傳 unknown）"""
    try:
        from streamlit.runtime.scriptrunner import get_script_run_ctx
        ctx = get_script_run_ctx()
        if ctx is None:
            return "unknown"
        from streamlit.runtime import get_instance
        client = get_instance().get_client(ctx.session_id)
        if client and hasattr(client, "request"):
            ip = (
                client.request.headers.get("X-Forwarded-For")
                or client.request.headers.get("X-Real-Ip")
                or getattr(client.request, "remote_ip", None)
                or "unknown"
            )
            return ip.split(",")[0].strip()
    except Exception:
        pass
    return "unknown"

def get_log():
    """
    取得帶有 [employee_id@client_IP] 前綴的 logger。
    在任何需要記錄使用者操作的地方呼叫：
        log = get_log()
        log.info("上傳音檔完成")
        log.warning("音檔為空")
        log.error("轉文字失敗", exc_info=True)   # ← 自動帶 traceback
    """
    import streamlit as _st
    user_id = _st.session_state.get("employee_id", "unknown")
    ip      = _get_client_ip()
    return get_user_logger(user_id, ip)
# ──────────────────────────────────────────────────────────────────────────────


import streamlit as st
import pandas as pd
import numpy as np
import time
import os, shutil
from app.qa.callcompliance import exceltodict, run_callcompliance
import io
from datetime import datetime, timedelta
import subprocess 
from app.utils.log import init_project_log_folder, append_log, write_and_protect, protect_file
from app.utils.open_localfile import local_path_to_download_button, local_path_to_http, local_path_to_http_url
import requests
import streamlit as st
import base64
import sys
import socket
import json
import streamlit.components.v1 as components
# from app.utils.lockfile import if_get_lock, set_lock_holder, if_release_lock_holder  # 已停用單人卡控
import app.qa.qa_agent as qa_agent
from app.qa.qa_agent import fmt_evidence, get_checkpoints, get_qa_result
from app.qa.load_question_sets import build_dynamic_qset
import google.generativeai as genai
from dotenv import load_dotenv
import os
from app.utils.load_history import load_history_from_logs
from app.utils.reset import reset_project
from app.utils.dedup_audio import dedup_by_name, _norm_path, _norm_paths
from app.utils.reupload_audio import _make_writable, _safe_remove, clean_project_media
from app.utils.utils import update_meta, parse_reuse_keywords_to_fileobjs, write_meta_and_header, unit_code_from_acceptance #,infer_unit_code_from_context
import io
from io import BytesIO
from app.audio.audio_duration import _read_bytes_preserve, get_audio_duration_seconds, format_duration, get_size_mb_from_bytes


st.set_page_config(layout="wide")

# ✅ 指定 .env 路徑
load_dotenv()
# ✅ 正確取得變數：只填變數名稱，不要填入完整路徑
VALID_USERNAME = os.getenv("VALID_USERNAME")
VALID_PASSWORD = os.getenv("VALID_PASSWORD")
print("✅ 環境變數讀取結果：", VALID_USERNAME, VALID_PASSWORD)

# ✅ Session flag: 是否登入成功
if "authenticated" not in st.session_state:
    st.session_state.authenticated = False

# ✅ 尚未登入時，顯示輸入框
if not st.session_state.authenticated:
    st.title("請登入音檔檢核系統")

    with st.form("login_form", clear_on_submit=True):
        username = st.text_input("使用者名稱")
        password = st.text_input("密碼", type="password")
        submitted = st.form_submit_button("登入")

        if submitted:
            if username == VALID_USERNAME and password == VALID_PASSWORD:
                st.success("✅ 登入成功，正在進入系統...")
                st.session_state.authenticated = True
                st.session_state.go_home = True  # ✅ 這會讓導航預設選「開新專案」
                time.sleep(1)
                st.rerun()  # 重新載入頁面進入主流程
            else:
                st.error("⛔ 帳號或密碼錯誤，請再試一次")

    st.stop()  # 停止後續所有內容，直到登入成功

# ✅ 已登入，載入歷史資料
if "authenticated" in st.session_state and st.session_state.authenticated:
    print(st.session_state.authenticated)

# 獲取本機IP
def get_local_ip():
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        # 嘗試連接到一個非內網 IP，不會真的發送資料，但可以取得正確的 local IP
        s.connect(("8.8.8.8", 80))
        local_ip = s.getsockname()[0]
    finally:
        s.close()
    return local_ip

LOCAL_IP = get_local_ip()
print(f"web啟動於IP：{LOCAL_IP}")

# Azure Whisper 不需載入本機模型

# 初始化流程步驟
if "step" not in st.session_state:
    st.session_state.step = 0
if "audio_folder" not in st.session_state:
    st.session_state.audio_folder = None
if "audio_upload_key" not in st.session_state:
    st.session_state.audio_upload_key = f"audio_upload_{time.time()}"
if "stt_done" not in st.session_state:
    st.session_state.stt_done = False
if "project_type" not in st.session_state:
    st.session_state.project_type = "開新專案"
if "last_project_type" not in st.session_state:
    st.session_state.last_project_type = None
if "clone_keyword_files" not in st.session_state:
    st.session_state.clone_keyword_files = []
if "unit_code" not in st.session_state:
    st.session_state.unit_code = '1'
if "user_split" not in st.session_state:
    st.session_state.user_split = None
if "user_split_confirme" not in st.session_state:
    st.session_state.user_split_confirme = False
if "step1_success" not in st.session_state:
    st.session_state.step1_success = False
if "step1_success_msg" not in st.session_state:
    st.session_state.step1_success_msg = ""

if "step1_success_clone" not in st.session_state:
    st.session_state.step1_success_clone = False
if "step1_success_msg_clone" not in st.session_state:
    st.session_state.step1_success_msg_clone = ""

valid_pages = ["開新專案", "歷史結果查看"]

# 編輯跳轉法1：處理 URL 參數
page_param = st.query_params.get("page", None)
default_page = page_param[0] if isinstance(page_param, list) else page_param or "開新專案"
if default_page not in valid_pages:
    default_page = "開新專案"

# 編輯跳轉法2：處理強制跳頁flag（最關鍵邏輯）
if st.session_state.get("force_page_jump"):
    st.session_state.project_type = "開新專案"
    default_page = "開新專案"
    del st.session_state["force_page_jump"]

default_index = valid_pages.index(default_page)

# 若點"回首頁"強制跳轉
if st.session_state.get("go_home"):
    st.session_state.project_type = "開新專案"
    default_page = "開新專案"
    del st.session_state["go_home"]

default_index = valid_pages.index(default_page)

# ➤ 1. 側邊欄功能導航
with st.sidebar:
    st.sidebar.markdown("<h3 style='font-size: 36px; font-weight: bold; margin-bottom:-20px;'> 🧭 功能導航</h3>", unsafe_allow_html=True)
    project_type = st.radio(
        "功能導航", 
        valid_pages, 
        index=default_index,
        key="project_type",         
        label_visibility="collapsed")
    print(f"radio 選擇project_type:{project_type}")

# 3) **先算 clone 與否，再做清理**（非常關鍵）
#    is_clone_mode_now 放在這裡：任何清理前先決定當輪是否為 clone
# =========================
is_clone = (st.session_state.get("reuse_mode") == "audio_and_transcript")
st.session_state["is_clone"] = is_clone  # 若其他地方要讀

# 若使用者功能切換時，清理 session（放在 radio 後面立刻做）
if st.session_state.project_type != st.session_state.get("last_project_type"):
    print("⚠️ 功能切換，清空相關狀態")
    st.session_state.history_query_triggered = False
    for key in [
        "acceptance_id",
        "log_header_written",
        "uploaded_audio_paths",
        "uploaded_keywords_path",
        "df_results",
        "show_results",
        # step1
        "step1_success",
        "step1_success_msg",
        "step1_success_clone",
        "step1_success_msg_clone",

        #step2
        "split_toggle",
        "start_stt"
        # step4
        "df_results_query_triggered",
        "last_df_filters",
        "success_msg",
        "df_results_info",
        "show_results",
        "log_result_path",
        "history",
        "df_results",
        "keywords_dict",
        "unit_code",
        "df_results_filter_confirmed",
        "excel_bytes_filtered",
        "export_filename_filtered",
        "export_path_filtered",
        "filtered_df_results_info",
        "show_filter",     # ← 必加
        # step_qa (AI質檢)
        "qa_done", "qa_results",
        "qa_show_filter",
        "qa_confirm_query", "qa_confirmed_status", "qa_confirmed_audio_ids",
        "qa_filter_confirmed", "qa_filter_excel_bytes", "qa_filter_export_filename",
        "qa_filter_export_path", "qa_filter_export_info",
        "qa_current_page", "qa_last_filter",
    ]:
        st.session_state.pop(key, None) #有key就刪，沒則不刪
    
    # 非 clone 才重置流程步驟，避免 clone 被誤清
    if not st.session_state["is_clone"]:
        st.session_state.step = 0
        st.session_state.stt_done = False

# 儲存當前功能為 last_project_type
st.session_state.last_project_type = st.session_state.project_type

# Spacer 撐高，讓底部區塊靠下
with st.sidebar:
    st.markdown("<div style='height: 400px;'></div>", unsafe_allow_html=True)

# 下方操作區塊
with st.sidebar:
    # st.markdown("### ⚙️ 系統操作")
    if st.button("🏠 回到頁首", help="重新開始整個流程"):
        reset_project()
        # 🧽 這邊加上 clone 模式的相關 flag 清除
        for key in [
            "log_header_written", 
            "clone_initialized", 
            "clone_file_objs",
            "remove_clone_indices",
            "run_remove_clone",

            "clone_kw_initialized",
            "clone_kw_initialized_for",
            "clone_keyword_files",
            "remove_clone_kw_indices",
            "run_remove_clone_kw",

            "clone_prefilled",
            # step1
            "step1_success",
            "step1_success_msg",
            "step1_success_clone",
            "step1_success_msg_clone",
            #step2
            "split_toggle",
            "start_stt"
            #step4
            "df_results_query_triggered","last_df_filters",
            "success_msg","df_results_info","show_results","log_result_path",
            "history","df_results","keywords_dict","unit_code",
            "last_df_filters","df_results_filter_confirmed",
            "excel_bytes_filtered","export_filename_filtered","export_path_filtered","filtered_df_results_info",
        ]:
            st.session_state.pop(key, None)
        st.success("✅ 流程已重新開始")
        st.session_state["go_home"] = True #flag
        st.rerun()  # 強制重整頁面以應用 reset

    # if st.button("↗️ 離開系統", help="清空目前工作階段並退出"):
    #     if "log_header_written" in st.session_state:
    #         del st.session_state["log_header_written"]
    #     # 直接清除 session，不再需要釋放 lockfile
    #     st.session_state.pop("lock_holder", None)
    #     st.session_state.pop("employee_id", None)
    #     st.success("✅ 使用者已離開系統")
    #     time.sleep(2)
    #     st.rerun()


    if st.session_state.get("is_clone", False):
        if st.button("🆕 回新專案", help="清空已載入的歷史資料，回到初始狀態"):
            reset_project()
            for key in [
                "reuse_mode",
                "is_clone",
                #clone 模式的相關 flag 清除
                "log_header_written", 
                "clone_initialized", 
                "clone_file_objs",
                "remove_clone_indices",
                "run_remove_clone",

                "clone_kw_initialized",
                "clone_kw_initialized_for",
                "clone_keyword_files",
                "remove_clone_kw_indices",
                "run_remove_clone_kw",

                "clone_prefilled",
                # step1
                "step1_success",
                "step1_success_msg",
                "step1_success_clone",
                "step1_success_msg_clone",
                #step2
                "split_toggle",
                "start_stt"
            ]:
                st.session_state.pop(key, None)
            st.success("✅ 已回開新專案")
            st.session_state["go_home"] = True #flag
            st.rerun()  # 強制重整頁面以應用 reset

# =========================
# 5) Step 0（開新專案頁）
# =========================
CODE_TO_UNIT = {"1": "服務中心", "2": "智能客服中心", "3": "直效行銷部"}
UNIT_TO_CODE = {v: k for k, v in CODE_TO_UNIT.items()}

if st.session_state.project_type == "開新專案":   
    st.markdown('<p style="font-size: 48px; font-weight: bold;">音檔檢核判讀 </p>', unsafe_allow_html=True)
    
    if st.session_state["is_clone"] and st.session_state.step >= 0:
        # ========== Step 0 Clone 模式 ==========    
        reuse_data = st.session_state.get("reuse_project_data") or {}
        src = reuse_data.get("原始受理編號") or reuse_data.get("acceptance_id")
        st.success(f"🔁 編輯歷史專案CLONE：{reuse_data['原始受理編號']}")#, 受理單位代碼：{reuse_data['受理單位代碼']}

        # 換了不同歷史案：重置 one-shot，並記下新的來源
        if src and st.session_state.get("clone_source") != src:
            st.session_state.clone_prefilled = False
            st.session_state.clone_source = src
            st.session_state.step = 0
            st.session_state.stt_done = False
            for k in [
                #step0
                "lock_holder",
                "acceptance_id","audio_folder","unit_code","unit_name","clone_prefilled",
                "audio_file_paths","transcript_paths",
                "uploaded_keywords_path","result_path",
                "log_header_written",
                #step3clone
                "clone_kw_initialized","clone_kw_initialized_for","clone_keyword_files",
                "remove_clone_kw_indices","run_remove_clone_kw",
                "log_keywords_path","uploaded_keywords",

                #step4
                "df_results_query_triggered","last_df_filters",
                "success_msg","df_results_info","show_results","log_result_path",
                "history","df_results","keywords_dict","unit_code",
                "last_df_filters","df_results_filter_confirmed",
                "excel_bytes_filtered","export_filename_filtered","export_path_filtered","filtered_df_results_info",
                # step_qa (AI質檢)
                "qa_done", "qa_results",
                "qa_show_filter",
                "qa_confirm_query", "qa_confirmed_status", "qa_confirmed_audio_ids",
                "qa_filter_confirmed", "qa_filter_excel_bytes", "qa_filter_export_filename",
                "qa_filter_export_path", "qa_filter_export_info",
                "qa_current_page", "qa_last_filter",
            ]:
                st.session_state.pop(k, None)            

        # ---------- one-shot：只預填，不畫 UI ----------
        if not st.session_state.get("clone_prefilled", False):
            st.session_state.employee_id = reuse_data.get("員工編號", "")
            unit_code_from_history = str(reuse_data.get("受理單位代碼", "4"))
            st.session_state.unit_code = unit_code_from_history
            print(f"st.session_state.unit_code: {st.session_state.unit_code}, type: {type(st.session_state.unit_code)}")
            st.session_state.unit_name = CODE_TO_UNIT.get(unit_code_from_history, "未知")
            st.session_state.clone_prefilled = True

        # ---------- Clone 的 UI（每輪都渲染；使用者可改值） ----------
        st.markdown('<p style="font-size: 36px; font-weight: bold;">Step 0. 建立專案基本資訊</p>', unsafe_allow_html=True)
        st.text_input("請輸入員工編號", key="employee_id")
        # --- 先確定員編 → 拿鎖 ---
        if st.button("確定員編"):
            emp = (st.session_state.get("employee_id") or "").strip()

            if not emp:
                st.warning("請先輸入員工編號")
            else:
                st.session_state["lock_holder"] = emp  # 保留 session key 相容性
                st.success(f"歡迎 {emp}，您已取得使用權。")

        # --- 拿到鎖後才顯示單位選擇 ---
        if st.session_state.get("lock_holder") and st.session_state.get("lock_holder") == st.session_state.get("employee_id"):
            
            components.html(f"""
            <script>
                const data = new Blob([JSON.stringify({{"current_user": "{st.session_state.get("employee_id")}"}})], {{ type: "application/json" }});
                window.addEventListener('beforeunload', function () {{
                    console.log("📤 current_user嘗試釋放鎖: {st.session_state.get("employee_id")}");
                    navigator.sendBeacon("http://{LOCAL_IP}:8505/release_lock", data);
                }});
            </script>
            """, height=1)#release_lock服務網址
            
            print(">>> 渲染前 unit_code:", st.session_state.get("unit_code"),"dtype:", type(st.session_state.unit_code))
            print(">>> options:", list(CODE_TO_UNIT.keys()))
            unit_code = st.session_state.get("unit_code", "1")
            st.session_state.unit_code = unit_code
            st.selectbox(
                "請選擇本次音檔單位",
                options=list(CODE_TO_UNIT.keys()),
                key="unit_code",
                format_func=lambda c: f"{c} {CODE_TO_UNIT[c]}",
            )

            if st.button("確定音檔單位"):
                # 產生受理編號與資料夾
                unit_code = st.session_state.unit_code  
                st.session_state.acceptance_id = f"{datetime.now():%Y%m%d_%H%M%S}_{unit_code}"
                st.session_state.step = 1
                st.session_state.stt_done = False 

                log_folder_path = init_project_log_folder(st.session_state.acceptance_id)
                st.session_state.audio_folder = log_folder_path

                # 寫meta.json
                meta_path = os.path.join(log_folder_path, "meta.json")
                with open(meta_path, "w", encoding="utf-8") as f:
                    json.dump({
                        "acceptance_id": st.session_state.acceptance_id,
                        "employee_id": st.session_state.employee_id,
                        "cloned_from": reuse_data.get("原始受理編號")
                    }, f, ensure_ascii=False, indent=2)
                
                # 寫log
                append_log(st.session_state.acceptance_id, f"員工編號：{st.session_state.employee_id}")
                append_log(st.session_state.acceptance_id, f"受理編號：{st.session_state.acceptance_id}")
                append_log(st.session_state.acceptance_id, f"Clone 自：{reuse_data.get('原始受理編號')}")

                #每次覆蓋 clone 音檔路徑
                st.session_state.audio_file_paths = reuse_data.get("音檔路徑清單", [])
                st.session_state.transcript_paths = reuse_data.get("逐字稿路徑清單")
                st.session_state.uploaded_keywords_path = reuse_data.get("關鍵字路徑")
                st.session_state.result_path = reuse_data.get("比對結果")

                print(f"音檔路徑清單:{reuse_data.get('音檔路徑清單')}") #皆檔案路徑、非資料夾
                print(f"逐字稿路徑清單:{reuse_data.get('逐字稿路徑清單')}")
                print(f"關鍵字路徑:{reuse_data.get('關鍵字路徑')}")
                print(f"比對結果:{reuse_data.get('比對結果')}")

        # Step 1～4 的共同標題區
        if "acceptance_id" in st.session_state:
            st.markdown(f"<p style='font-size:18px;'>🆔 新受理編號：{st.session_state.acceptance_id}</p>", unsafe_allow_html=True)    

    # ========== Step 0 非 Clone 模式 ==========    
    elif (not st.session_state.get("is_clone", False)) and st.session_state.step >= 0:
        st.markdown('<p style="font-size: 36px; font-weight: bold;">Step 0. 建立專案基本資訊</p>', unsafe_allow_html=True)
        
        # --- 1) 員編輸入（用 key 綁定，值留在 session）---
        st.text_input("請輸入員工編號", key="employee_id", label_visibility="visible")

        # --- 2) 確定員編 ---
        if st.button("確定員編"):
            emp = (st.session_state.get("employee_id") or "").strip()

            if not emp:
                st.warning("請先輸入員工編號")
            else:
                st.session_state["lock_holder"] = emp  # 保留 session key 相容性
                st.success(f"歡迎 {emp}，您已取得使用權。")

                # ✅ 這個動作只是改員編，不改單位。
                # 若已經有受理編號，讓畫面上的 unit_code 直接同步成「受理編號尾碼」，
                # 如此 selectbox 會維持原來的單位，不會跳回 1。
                if "acceptance_id" in st.session_state:
                    st.session_state.unit_code = unit_code_from_acceptance()  # ← 關鍵同步
                    write_meta_and_header(force_update_log=True)
                    st.info(f"已更新員編，單位不變（{CODE_TO_UNIT[st.session_state.unit_code]}），受理編號維持不變。")

        # --- 3) 單位選擇（只有拿到鎖的人可以操作）---
        if st.session_state.get("lock_holder") == st.session_state.get("employee_id"):

            components.html(f"""
            <script>
                const data = new Blob([JSON.stringify({{"current_user": "{st.session_state.get("employee_id")}"}})], {{ type: "application/json" }});
                window.addEventListener('beforeunload', function () {{
                    console.log("📤 current_user嘗試釋放鎖: {st.session_state.get("employee_id")}");
                    navigator.sendBeacon("http://{LOCAL_IP}:8505/release_lock", data);
                }});
            </script>
            """, height=1)#release_lock服務網址height=0, width=0, scrolling=False

            # 在建立 selectbox 之前，若還沒有 unit_code，或 acceptance_id 存在但兩者不一致，
            # 先以受理編號尾碼「校正」一次（避免顯示錯誤的預設 1）。
            acc_code = unit_code_from_acceptance()
            if "unit_code" not in st.session_state:
                st.session_state.unit_code = acc_code
            elif "acceptance_id" in st.session_state and st.session_state.unit_code not in CODE_TO_UNIT:
                st.session_state.unit_code = acc_code

            # 建立 selectbox：只用 key，**不傳 index / value** → 不會跳黃色警告
            st.selectbox(
                "請選擇本次音檔單位",
                options=list(CODE_TO_UNIT.keys()),
                key="unit_code",
                format_func=lambda c: f"{c} {CODE_TO_UNIT[c]}",
            )

            # 使用者想改單位才按這顆按鈕
            if st.button("確定音檔單位"):
                new_code = st.session_state.unit_code
                curr_acc = st.session_state.get("acceptance_id")
                curr_code = curr_acc.rsplit("_", 1)[-1] if curr_acc else None

                if curr_acc and curr_code == new_code:
                    # 單位沒變：受理編號不變，只覆寫 meta ＆ 記更新
                    write_meta_and_header(force_update_log=True)
                    st.info("已更新員編，單位不變，受理編號維持不變。")
                else:
                    # 單位變或尚未建立：建立新受理編號 + 新資料夾，寫 header
                    for k in [
                        "audio_folder", "audio_file_paths", "audio_file_names",
                        "transcript_paths", "uploaded_keywords_path", "result_path",
                        "df_results", "show_results"
                    ]:
                        st.session_state.pop(k, None)

                    new_acc = f"{datetime.now():%Y%m%d_%H%M%S}_{new_code}"
                    st.session_state.acceptance_id = new_acc
                    st.session_state.step = max(st.session_state.step, 1)
                    st.session_state.stt_done = False
                    st.session_state.audio_upload_key = f"audio_upload_{time.time()}"

                    write_meta_and_header(force_update_log=False)
                    st.success(f"已建立新專案：{new_acc}")
            
        # --- 4) 顯示目前受理編號 ---
        if "acceptance_id" in st.session_state:
            st.markdown(f"<p style='font-size:18px;'>🆔 受理編號： {st.session_state.acceptance_id}</p>", unsafe_allow_html=True)

    # ========== 頁籤介面：音檔處理 & 關鍵字比對 ==========
    # 只有在 Step 0 完成後（有 acceptance_id）才顯示頁籤
    if "acceptance_id" in st.session_state:
        # 自訂頁籤樣式：36px 字體大小和粗體
        st.markdown("""
        <style>
        /* 頁籤按鈕樣式 */
        .stTabs [data-baseweb="tab-list"] button [data-testid="stMarkdownContainer"] p {
            font-size: 18px;
            font-weight: bold;
        }
        /* 頁籤按鈕本身 */
        .stTabs [data-baseweb="tab-list"] button {
            font-size: 36px;

        }
        </style>
        """, unsafe_allow_html=True)
        
        tab1, tab2, tab3 = st.tabs(["📁 產出逐字稿", "🔍 比對關鍵字", "👾 AI 質檢"])
        
        # ==================== 頁籤一：音檔處理 ====================
        with tab1:
            st.markdown('<p style="font-size: 36px; font-weight: bold;">Step 1. 檢核聲音檔案 - 上傳區</p>', unsafe_allow_html=True)
            
            # ========== Step 1 Clone ==========
            if st.session_state["is_clone"]:
                # 先確保有控制widget token
                if "uploader_token" not in st.session_state:
                    st.session_state.uploader_token = f"audio_upload_{int(time.time()*1000)}" #控制鍵

                # 1 file_uploader 讓使用者額外加檔
                uploaded_files = st.file_uploader(
                    "請上傳音檔（支援 MP3, MP4, WAV, M4A；上限200MB）",
                    type=["mp3", "mp4", "wav", "m4a"],
                    accept_multiple_files=True,
                    key=st.session_state.uploader_token, #用控制變數當真正的 widget key
                )

                # 2 每次編輯都會重新載入歷史專案
                reuse_data = st.session_state.get("reuse_project_data", {})
                audio_paths = reuse_data.get("音檔路徑清單", [])

                # 初始化 clone 模式的檔案物件
                if not st.session_state.get("clone_initialized", False):

                    st.session_state.clone_file_objs = []
                    for path in audio_paths:
                        if path and os.path.exists(path) and not path.lower().endswith(("_left.wav", "_right.wav")):
                            with open(path, "rb") as f:
                                file_obj = io.BytesIO(f.read())
                                file_obj.name = os.path.basename(path)
                                st.session_state.clone_file_objs.append(file_obj)
                                print(f"✅ clone_file_obj: {file_obj.name}, {len(file_obj.getbuffer()) / 1024 / 1024:.1f} MB")
                    st.session_state.clone_initialized = True

                # 3 初始化移除用的 session_state
                if "remove_clone_indices" not in st.session_state:
                    st.session_state.remove_clone_indices = set()
                if "run_remove_clone" not in st.session_state:
                    st.session_state.run_remove_clone = False

                # 如果有移除指令就先執行
                if st.session_state.run_remove_clone:

                    to_remove = set(st.session_state.remove_clone_indices)  # 已是 set，但保險再包一次
                    st.session_state.clone_file_objs = [
                        obj for i, obj in enumerate(st.session_state.clone_file_objs)
                        if i not in to_remove
                    ]

                    st.session_state.remove_clone_indices.clear()
                    st.session_state.run_remove_clone = False

                # 顯示 clone 檔案清單與移除按鈕
                st.markdown("🔁 載入歷史主音檔：")
                for idx, f in enumerate(st.session_state.clone_file_objs):
                    col1, col2 = st.columns([8, 1])
                    with col1:
                        st.markdown(f"📄 **{f.name}** ({len(f.getbuffer()) / 1024 / 1024:.1f} MB)")
                    with col2:
                        if st.button("❌", key=f"remove_clone_{idx}"):
                            st.session_state.remove_clone_indices.add(idx)
                            st.session_state.run_remove_clone = True
                            st.rerun()  # 🔁 強制重新執行（立即反映）
                        
                # ---------------- 顯示「已選擇檔案」區（合併 clone + 剛上傳） ----------------
                # 先把 uploaded_files 轉成 BytesIO（方便重複讀）
                processed_uploaded = []
                if uploaded_files:
                    for uf in uploaded_files:
                        raw = uf.read()
                        bio = BytesIO(raw)
                        bio.name = getattr(uf, "name", "unnamed")
                        processed_uploaded.append(bio)

                # 合併要給使用者看的清單：先 clone，再新上傳
                combined_for_display = []
                combined_for_display.extend(st.session_state.clone_file_objs or [])
                combined_for_display.extend(processed_uploaded or [])

                if combined_for_display:
                    # 計算總大小與總時長
                    total_seconds = 0.0
                    total_size_mb = 0.0
                    st.markdown(f"**已選 {len(combined_for_display)} 個音檔：**")
                    for obj in combined_for_display:
                        # obj 可能是 BytesIO（clone）或剛建立的 BytesIO（uploaded）
                        # 取 bytes 與大小
                        try:
                            data_bytes = obj.getvalue()
                        except Exception:
                            try:
                                obj.seek(0)
                                data_bytes = obj.read()
                            except Exception:
                                data_bytes = b""
                        size_mb = len(data_bytes) / 1024 / 1024
                        dur = get_audio_duration_seconds(BytesIO(data_bytes)) or 0.0
                        total_size_mb += size_mb
                        total_seconds += dur
                        st.markdown(f"• **{getattr(obj, 'name', 'unnamed')}** ({size_mb:.1f} MB · {format_duration(dur)})")

                    # 顯示總計
                    st.markdown(f"**總計大小 {total_size_mb:.1f} MB · 總計長度 {format_duration(total_seconds)}**")
                else:
                    st.info("尚未選擇音檔。請點 Browse files 或拖曳上傳音檔。")


                # ─── 2) 全寬 banner 區（放在整個頁面最上方，所有 Step 區塊之前） ───
                banner_slot = st.empty()  # 這就是你要的「全平 placeholder」
                if st.session_state.get("step1_success_clone"):
                    banner_slot.success(st.session_state.step1_success_msg_clone)

                # 4 兩顆按鈕（左：重新上傳、右：確定使用）
                msg_ph = st.empty() 
                left, right = st.columns([1, 1])
                with left:
                    if st.button("確定使用音檔", disabled=st.session_state.get("step1_success_clone", False)):
                        # (1) 合併檔案並去重
                        combined = []
                        # clone_file_objs 已是 BytesIO，可直接 extend
                        if st.session_state.clone_file_objs:
                            combined.extend(st.session_state.clone_file_objs)
                        # processed_uploaded 是 BytesIO
                        if processed_uploaded:
                            combined.extend(processed_uploaded)
                        final_files = dedup_by_name(combined)

                        if not final_files:
                            msg_ph.warning("請至少上傳一個音檔。")
                        else:
                            # (2) 寫到專案資料夾
                            log_folder_path = init_project_log_folder(st.session_state.acceptance_id)
                            st.session_state.audio_folder = log_folder_path

                            saved_paths = []
                            for fobj in final_files:
                                dst = os.path.join(log_folder_path, getattr(fobj, "name", "unnamed"))
                                # 取 bytes（適用 BytesIO 或 UploadedFile）
                                try:
                                    # BytesIO: getvalue(), UploadedFile: read()
                                    data_bytes = fobj.getvalue()
                                except Exception:
                                    try:
                                        fobj.seek(0)
                                        data_bytes = fobj.read()
                                    except Exception:
                                        data_bytes = b""
                                write_and_protect(dst, data_bytes)
                                append_log(st.session_state.acceptance_id, f"上傳音檔：{dst}")
                                saved_paths.append(dst)

                            # (3) 用統一命名寫回 session（Step 2 只看這兩個鍵）
                            st.session_state.audio_file_paths = saved_paths
                            st.session_state.audio_file_names = [os.path.basename(p) for p in saved_paths]

                            # (4) 寫入 meta.json
                            meta_path = os.path.join(log_folder_path, "meta.json")
                            try:
                                meta = {}
                                if os.path.exists(meta_path):
                                    with open(meta_path, "r", encoding="utf-8") as f:
                                        meta = json.load(f)
                                meta["audio_names"] = [os.path.basename(p) for p in saved_paths]
                                meta["audio_paths"] = saved_paths
                                with open(meta_path, "w", encoding="utf-8") as f:
                                    json.dump(meta, f, ensure_ascii=False, indent=2)
                            except Exception as e:
                                st.warning(f"⚠️ 無法寫入 meta.json：{e}")

                            # (5) 顯示成功訊息
                            total_saved_seconds = 0.0
                            total_saved_size = 0.0
                            file_lines = []
                            for p in saved_paths:
                                try:
                                    with open(p, "rb") as fh:
                                        b = fh.read()
                                        dur = get_audio_duration_seconds(BytesIO(b)) or 0.0
                                        size_mb = len(b) / 1024 / 1024
                                except Exception:
                                    dur = 0.0
                                    size_mb = 0.0
                                total_saved_seconds += dur
                                total_saved_size += size_mb
                                fname = os.path.basename(p)
                                file_lines.append(f"• {fname} ({format_duration(dur)})")
                            file_names = "  \n".join(file_lines)

                            st.session_state.step1_success_clone = True
                            st.session_state.step1_success_msg_clone = (
                                f"✅ 已上傳 {len(saved_paths)} 個音檔（總計大小 {total_saved_size:.1f} MB · 總計長度: {format_duration(total_saved_seconds)}）：\n\n{file_names}"
                            )

                            # (6) 標記完成，不需要改 step
                            st.session_state.stt_done = False
                            st.rerun()
                with right:
                    if st.button("重新上傳", help="清空目前已選與已存音檔，重新選擇"):
                        # 只有「重新上傳」會清除 banner
                        st.session_state.step1_success_clone = False
                        st.session_state.step1_success_msg_clone = ""

                        # (1) 刪除專案資料夾內音檔
                        audio_folder = st.session_state.get("audio_folder")
                        if audio_folder:
                            clean_project_media(
                                project_folder=audio_folder,
                                log_id=st.session_state.acceptance_id
                            )
                        # (2) 寫入 meta.json
                        meta_path = os.path.join(audio_folder, "meta.json")
                        try:
                            meta = {}
                            if os.path.exists(meta_path):
                                with open(meta_path, "r", encoding="utf-8") as f:
                                    meta = json.load(f)
                            meta["audio_names"] = None
                            meta["audio_paths"] = None
                            with open(meta_path, "w", encoding="utf-8") as f:
                                json.dump(meta, f, ensure_ascii=False, indent=2)
                        except Exception as e:
                            msg_ph.warning(f"⚠️ 無法寫入 meta.json：{e}")
                            
                        # (3) 清空 session 旗標選擇
                        for k in ["audio_file_paths",
                                  "audio_file_names",
                                  "stt_done",
                                  "step1_success_msg_clone"
                                  ]:
                            st.session_state.pop(k, None)
                        st.session_state.stt_btn_disabled = False
                        # 清 clone 暫存
                        st.session_state.clone_initialized = False
                        st.session_state.clone_file_objs = []
                        st.session_state.remove_clone_indices = set()
                        st.session_state.run_remove_clone = False
                        
                        # (4) 重新上傳：換 token（而不是動到同名 widget key），再 rerun
                        old_key = st.session_state.uploader_token
                        st.session_state.uploader_token = f"audio_upload_{int(time.time()*1000)}"
                        st.session_state.pop(old_key, None)   # 可選：移除舊 widget 的值
                        
                        # (5) 立即重繪，回到 Browse files 空白
                        st.rerun()
                    


            # ========== Step 1 非 Clone ==========
            else:  # not is_clone
                if "uploader_token" not in st.session_state:
                    st.session_state.uploader_token = f"audio_upload_{int(time.time()*1000)}"

                # 1. 上傳音檔區
                uploaded_files = st.file_uploader(
                    "請上傳音檔（支援 MP3, MP4, WAV, M4A；上限200MB）",
                    type=["mp3", "mp4", "wav", "m4a"],
                    accept_multiple_files=True,  # 允許多檔上傳
                    key=st.session_state.uploader_token,
                )

                # 3) 立即解析已選檔案（轉成 BytesIO 保存，方便重複讀取）
                processed_uploaded_files = []  # list of BytesIO objects with .name
                if uploaded_files:
                    for uf in uploaded_files:
                        # 讀一次 bytes，建立 BytesIO（之後可以重複用）
                        raw = uf.read()
                        bio = BytesIO(raw)
                        bio.name = getattr(uf, "name", "unnamed")
                        processed_uploaded_files.append(bio)

                # 4) 在 UI 顯示已選檔案的大小與時長，並計算總時長
                if processed_uploaded_files:
                    st.markdown(f"**已選 {len(processed_uploaded_files)} 個音檔：**")
                    total_seconds = 0.0
                    total_size = 0.0
                    for bio in processed_uploaded_files:
                        dur = get_audio_duration_seconds(bio)   # 可能為 None
                        dur_for_sum = dur if (dur is not None) else 0.0
                        total_seconds += dur_for_sum
                        size_mb = get_size_mb_from_bytes(bio.getvalue())
                        size_for_sum = size_mb if (size_mb is not None) else 0.0
                        total_size += size_for_sum
                        st.markdown(f"• **{bio.name}** ({size_mb:.1f} MB · {format_duration(dur)})")
                    # 顯示總計長度（醒目）
                    st.markdown(f"**總計大小 {total_size:.1f} MB · 總計長度 {format_duration(total_seconds)}**")

                else:
                    st.info("尚未選擇音檔。請點 Browse files 或拖曳上傳音檔。")

                # ─── 2) 全寬 banner 區（放在整個頁面最上方，所有 Step 區塊之前） ───
                banner_slot = st.empty()  # 這就是你要的「全平 placeholder」
                if st.session_state.get("step1_success"):
                    banner_slot.success(st.session_state.step1_success_msg)

                # 4 兩顆按鈕（左：重新上傳、右：確定使用）
                msg_ph = st.empty() 
                left, right = st.columns([1, 1])
                with left:
                    if st.button("確定使用音檔", disabled=st.session_state.get("step1_success", False)):
                        if not processed_uploaded_files:
                            msg_ph.warning("請至少上傳一個音檔。")
                        else:
                            # (1) 上傳檔案並去重
                            final_files = dedup_by_name(processed_uploaded_files)
                            # (2) 寫到專案資料夾
                            log_folder_path = init_project_log_folder(st.session_state.acceptance_id)
                            st.session_state.audio_folder = log_folder_path

                            saved_paths = []
                            for f in final_files:
                                dst = os.path.join(log_folder_path, f.name)
                                write_and_protect(dst, f.read())
                                append_log(st.session_state.acceptance_id, f"上傳音檔：{dst}")
                                saved_paths.append(dst)

                            # (3) 用統一命名寫回 session
                            st.session_state.audio_file_paths = saved_paths             # ✅ 之後所有流程都用這個
                            st.session_state.audio_file_names = [os.path.basename(p) for p in saved_paths]

                            # (4) 寫入 meta.json
                            meta_path = os.path.join(log_folder_path, "meta.json")
                            try:
                                meta = {}
                                if os.path.exists(meta_path):
                                    with open(meta_path, "r", encoding="utf-8") as f:
                                        meta = json.load(f)
                                meta["audio_names"] = [os.path.basename(p) for p in saved_paths]
                                meta["audio_paths"] = saved_paths
                                with open(meta_path, "w", encoding="utf-8") as f:
                                    json.dump(meta, f, ensure_ascii=False, indent=2)
                            except Exception as e:
                                msg_ph.warning(f"⚠️ 無法寫入 meta.json：{e}")

                            # (5) 顯示成功訊息（同時計算已儲存檔案的總時長）
                            total_saved_seconds = 0.0
                            file_lines = []
                            for p in saved_paths:
                                try:
                                    with open(p, "rb") as fh:
                                        dur = get_audio_duration_seconds(fh) or 0.0
                                except Exception:
                                    dur = 0.0
                                total_saved_seconds += dur
                                fname = os.path.basename(p)
                                file_lines.append(f"• {fname} ({format_duration(dur)})")
                            file_names = "  \n".join(file_lines)

                            st.session_state.step1_success = True
                            st.session_state.step1_success_msg = (
                                f"✅ 已上傳 {len(saved_paths)} 個音檔（總計大小 {total_size:.1f} MB · 總計長度: {format_duration(total_saved_seconds)}）：\n\n{file_names}"
                            )
                            # (6) 標記完成，不需要改 step
                            st.session_state.stt_done = False
                            st.session_state.stt_btn_disabled = False
                            st.rerun()
                    
                with right:
                    if st.button("重新上傳", help="清空目前已存/已選音檔，回到空白"):
                        # 只有「重新上傳」會清除 banner
                        st.session_state.step1_success = False
                        st.session_state.step1_success_msg = ""

                        # (1) 刪除專案資料夾內音檔
                        audio_folder = st.session_state.get("audio_folder")
                        if audio_folder:
                            clean_project_media(
                                project_folder=audio_folder,
                                log_id=st.session_state.acceptance_id
                            )

                        # (2) 寫入 meta.json
                        meta_path = os.path.join(audio_folder, "meta.json")
                        try:
                            meta = {}
                            if os.path.exists(meta_path):
                                with open(meta_path, "r", encoding="utf-8") as f:
                                    meta = json.load(f)
                            meta["audio_names"] = None
                            meta["audio_paths"] = None
                            with open(meta_path, "w", encoding="utf-8") as f:
                                json.dump(meta, f, ensure_ascii=False, indent=2)
                        except Exception as e:
                            msg_ph.warning(f"⚠️ 無法寫入 meta.json：{e}")
                            
                        # (3) 清空目前選擇
                        for k in ["audio_file_paths", 
                                  "audio_file_names",
                                  "stt_done",
                                  "transcript_paths"
                                  ]:
                            st.session_state.pop(k, None)
                        st.session_state.stt_btn_disabled = False

                        # (4) 重置 uploader（換 key）
                        old_key = st.session_state.uploader_token
                        st.session_state.uploader_token = f"audio_upload_{int(time.time()*1000)}"
                        st.session_state.pop(old_key, None)  # 可選，丟掉舊 widget 的值
                        msg_ph.empty()

                        # (5) 立即重繪，回到 Browse files 空白
                        st.rerun()

            # ========== Step 2: 語音轉文字（STT） ==========
            st.markdown('----------------------------------')
            st.markdown('<p style="font-size: 36px; font-weight: bold;">Step 2. 語音轉文字（STT）</p>', unsafe_allow_html=True)

            
            # 舊專案資訊（可能為空）
            reuse_data = st.session_state.get("reuse_project_data", {})
            reuse_main_names = [os.path.basename(p) for p in reuse_data.get("音檔路徑清單", [])
                                if not (p.endswith("_left.wav") or p.endswith("_right.wav"))]
            original_user_split = reuse_data.get("切角色模式", "unknown")

            # 目前實際使用的主檔名清單（Step1 已寫入 session）
            current_names = st.session_state.get("audio_file_names", [])
            
            # 建立逐字稿子資料夾
            project_folder = st.session_state.audio_folder
            transcript_folder = os.path.join(project_folder, "逐字稿")
            os.makedirs(transcript_folder, exist_ok=True)

            user_split = st.toggle("是否自動分音軌", value=False, key="split_toggle")

            if "stt_btn_disabled" not in st.session_state:
                st.session_state.stt_btn_disabled = False

            def disable_stt_btn():
                st.session_state.stt_btn_disabled = True

            # 若主檔名清單 + 切軌模式完全一致->沿用舊逐字稿
            if st.button("開始轉檔", key="start_stt", disabled=st.session_state.stt_btn_disabled, on_click=disable_stt_btn):
                
                # 條件符合：沿用舊逐字稿
                print(f"current_names:{current_names} vs reuse_main_names:{reuse_main_names}")
                print(f"current_user_split:{user_split} vs original_user_split:{original_user_split}")
                print(f"reuse_data.get(逐字稿路徑清單):{reuse_data.get('逐字稿路徑清單')}")            
                if (current_names == reuse_main_names) and (user_split == original_user_split) and reuse_data.get('逐字稿路徑清單'):
                    old_transcripts = reuse_data.get('逐字稿路徑清單', [])
                    dest_paths = []

                    for src in old_transcripts:
                        if not src:
                            continue
                        
                        # 智能路徑解析：處理舊專案（根目錄）和新專案（子資料夾）
                        actual_src = None
                        
                        # 1. 嘗試使用原始路徑
                        if os.path.isfile(src):
                            actual_src = src
                        else:
                            # 2. 如果原始路徑不存在，嘗試在逐字稿子資料夾中尋找
                            # 提取原專案資料夾路徑（假設格式為 D:\web\logs\YYYYMMDD_HHMMSS_N\...）
                            src_parts = src.split(os.sep)
                            if 'logs' in src_parts:
                                logs_idx = src_parts.index('logs')
                                if logs_idx + 1 < len(src_parts):
                                    # 原專案資料夾名稱（例如：20260212_172824_1）
                                    old_project_folder = os.sep.join(src_parts[:logs_idx+2])
                                    # 檔案名稱
                                    filename = os.path.basename(src)
                                    # 嘗試在逐字稿子資料夾中尋找
                                    alt_src = os.path.join(old_project_folder, "逐字稿", filename)
                                    if os.path.isfile(alt_src):
                                        actual_src = alt_src
                        
                        # 如果找到檔案，複製到新專案
                        if actual_src:
                            dst = os.path.join(transcript_folder, os.path.basename(actual_src))
                            try:
                                shutil.copy2(actual_src, dst)  # 保留時間戳等中繼資料
                                dest_paths.append(dst)
                                print(f"✅ 成功複製逐字稿：{actual_src} -> {dst}")
                            except Exception as e:
                                st.warning(f"⚠️ 無法複製逐字稿：{actual_src} -> {e}")
                        else:
                            print(f"⚠️ 找不到逐字稿檔案：{src}")

                    if dest_paths:
                        #  更新 Session（統一用複數，並保留單數相容）
                        st.session_state.stt_done = True
                        st.session_state.transcript_paths = dest_paths  # 存成 list

                        #  寫入 meta.json（存放在專案根目錄）
                        meta_path = os.path.join(project_folder, "meta.json")
                        try:
                            meta = {}
                            if os.path.exists(meta_path):
                                with open(meta_path, "r", encoding="utf-8") as f:
                                    meta = json.load(f)
                            meta["use_speaker_split"] = bool(user_split)
                            meta["transcript_source"] = "reuse"
                            meta["transcript_files"] = [os.path.basename(p) for p in dest_paths]
                            with open(meta_path, "w", encoding="utf-8") as f:
                                json.dump(meta, f, ensure_ascii=False, indent=2)
                        except Exception as e:
                            st.warning(f"⚠️ 無法寫入 meta.json：{e}")

                        # 寫入log
                        append_log(st.session_state.acceptance_id, f"沿用逐字稿：{src} -> {dst}")
                        append_log(st.session_state.acceptance_id, f"沿用逐字稿完成，共複製 {len(dest_paths)} 個檔案至 {transcript_folder}")
                        copied_list = "  \n".join([f"• {os.path.basename(p)}" for p in dest_paths])
                        st.success(f"📝 沿用原逐字稿，已複製到新專案：  \n{copied_list}")
                        st.session_state.stt_btn_disabled = False
                        st.rerun()
                    else:
                        st.warning("⚠️ 找到條件相符，但舊專案中沒有可複製的逐字稿檔案。")
                        st.session_state.stt_btn_disabled = False
                        st.rerun()

                #重轉逐字稿
                else:
                    if not st.session_state.stt_done:
                        # 1) 用已落地檔案路徑
                        disk_paths = st.session_state.get("audio_file_paths", [])
                        if not disk_paths:
                            st.error("找不到要上傳的音檔路徑，請回到 Step 1 重新確認。")
                            st.stop()

                        input_folder = st.session_state.audio_folder  # 直接用 Step1 落地資料夾

                        # 🔍 查詢 GPU 佇列狀態，顯示詳細排隊資訊
                        from app.audio.transcribe import is_lock_busy, get_queue_snapshot
                        import uuid as _uuid

                        # 產生本次任務的唯一 ID，傳給 transcribe_folder 讓它加入佇列
                        _my_job_id = str(_uuid.uuid4())

                        _gpu_busy = is_lock_busy()
                        
                        # ── 取得目前佇列中排在我前面的任務 (僅在 GPU 忙碌時) ──
                        _queue = get_queue_snapshot() if _gpu_busy else []

                        # ── 取得我自己的音檔清單與時長 ──
                        # 若有 split，每個原始音檔會產生 原檔 + left + right，共 3 個轉錄目標
                        _my_file_paths = st.session_state.get("audio_file_paths", [])
                        _my_files = []
                        for _p in _my_file_paths:
                            try:
                                with open(_p, "rb") as _fh:
                                    _dur = get_audio_duration_seconds(BytesIO(_fh.read())) or 0.0
                            except Exception:
                                _dur = 0.0
                            _fname = os.path.basename(_p)
                            _base, _ext = os.path.splitext(_fname)
                            if user_split:
                                # 切軌後會有原檔 + left + right，各自時長相同
                                _my_files.append({"name": _fname,              "duration_sec": _dur})
                                _my_files.append({"name": f"{_base}_left{_ext}",  "duration_sec": _dur})
                                _my_files.append({"name": f"{_base}_right{_ext}", "duration_sec": _dur})
                            else:
                                _my_files.append({"name": _fname, "duration_sec": _dur})

                        # ── 計算總時長 ──
                        _queue_total_sec = sum(
                            fi["duration_sec"]
                            for job in _queue for fi in job["files"]
                        )
                        _my_total_sec = sum(f["duration_sec"] for f in _my_files)
                        _grand_total_sec = _queue_total_sec + _my_total_sec
                        _estimated_sec = _grand_total_sec / 6  # Whisper Azureapi 約為音檔時長 1/6

                        # ── 顯示資訊卡片 ──
                        _row = 1
                        if _gpu_busy:
                            st.warning("⏳ GPU 忙碌中，您已進入排隊...")
                            if _queue:
                                st.markdown("##### 📋 在您之前的任務：")
                                for _job in _queue:
                                    for _fi in _job["files"]:
                                        st.markdown(
                                            f"&nbsp;&nbsp;**{_row}.** `{_fi['name']}` &nbsp;｜&nbsp; "
                                            f"{format_duration(_fi['duration_sec'])}",
                                            unsafe_allow_html=True
                                        )
                                        _row += 1
                                st.markdown("---")
                        
                        _split_note = "（已展開原檔 + left + right）" if user_split else ""
                        st.markdown(f"##### 📁 您的任務{_split_note}：")
                        for _i, _fi in enumerate(_my_files, _row):
                            st.markdown(
                                f"&nbsp;&nbsp;**{_i}.** `{_fi['name']}` &nbsp;｜&nbsp; "
                                f"{format_duration(_fi['duration_sec'])}",
                                unsafe_allow_html=True
                            )
                        st.markdown("---")
                        
                        st.info(
                            f"📊 **總計音檔長度：{format_duration(_grand_total_sec)}**　　"
                            f"⏱️ **預估轉錄時間：約 {format_duration(_estimated_sec)}**"
                        )


                        with st.spinner("🎙️ 正在呼叫 Azure Whisper 轉文字，請稍候..."):
                            try:
                                # 1) 切軌（可選）
                                produced_wavs = []
                                if bool(user_split):
                                    from app.audio.split import split_folder
                                    produced_wavs = split_folder(input_folder)

                                # 2) 呼叫 Azure Whisper API 轉文字
                                from app.audio.transcribe import transcribe_folder
                                produced_txts = transcribe_folder(
                                    input_folder=input_folder,
                                    output_dir=transcript_folder,
                                    model_size="azureapi",
                                    job_id=_my_job_id,
                                )

                                saved_paths = sorted(produced_wavs + produced_txts)
                                saved_names = [os.path.basename(p) for p in saved_paths]

                                
                                # 成功才更新 Session（統一使用複數 + 保留單數相容）
                                if saved_paths:
                                    st.session_state.stt_done = True
                                    st.session_state.transcript_paths = saved_paths

                                    # 寫入 meta.json（存放在專案根目錄）
                                    meta_path = os.path.join(project_folder, "meta.json")
                                    try:
                                        meta = {}
                                        if os.path.exists(meta_path):
                                            with open(meta_path, "r", encoding="utf-8") as f:
                                                meta = json.load(f)
                                        meta["use_speaker_split"] = bool(user_split)
                                        meta["transcript_source"] = "azure"                 # 來源：Azure Whisper API
                                        meta["transcript_files"] = saved_names            # 逐字稿檔名清單
                                        with open(meta_path, "w", encoding="utf-8") as f:
                                            json.dump(meta, f, ensure_ascii=False, indent=2)
                                    except Exception as e:
                                        st.warning(f"⚠️ 無法寫入 meta.json：{e}")
                                    # 寫入 log
                                    append_log(st.session_state.acceptance_id, f"GPU 轉文字完成，共儲存 {len(saved_paths)} 個音檔/逐字稿至 {transcript_folder}")
                                    stt_list = "  \n".join([f"• {os.path.basename(p)}" for p in saved_paths])
                                    st.success(f"📝 轉檔完成，儲存 {len(saved_paths)} 個檔案  \n{stt_list}")                             
                                    st.session_state.stt_btn_disabled = False
                                    st.rerun()
                                else:
                                    st.warning("⚠️ 這次沒有產生任何新的逐字稿或切軌檔。")
                                    st.session_state.stt_btn_disabled = False
                                    st.rerun()
                            except Exception as e:
                                st.error(f"❌ 轉檔流程錯誤：{e}")
                                st.session_state.stt_btn_disabled = False
                                st.rerun()
                        
            if st.session_state.stt_done:
                st.markdown('<p style="font-size: 24px; font-weight: bold;">逐字稿路徑</p>', unsafe_allow_html=True)
                _tp = st.session_state.get("transcript_paths", [])
                if _tp:
                    # 顯示個別逐字稿檔案路徑
                    _tp_rows = [{
                        "逐字稿路徑": p,
                        "轉檔時間": datetime.now().strftime("%Y/%m/%d %H:%M:%S")
                    } for p in _tp]
                    st.table(pd.DataFrame(_tp_rows))
                else:
                    # fallback：只顯示資料夾
                    st.table(pd.DataFrame([{
                        "逐字稿路徑": transcript_folder,
                        "轉檔時間": datetime.now().strftime("%Y/%m/%d %H:%M:%S")
                    }]))
                st.info("✅ 逐字稿已完成，請切換到「頁籤二：比對關鍵字」、 「頁籤三：AI 質檢」繼續操作")

        # ==================== 頁籤二：關鍵字比對 ====================
        with tab2:
            st.markdown('<p style="font-size: 36px; font-weight: bold;">上傳關鍵字</p>', unsafe_allow_html=True)
            
            # ── 防呆輔助函式 ──
            def _validate_keywords(f):
                """回傳 (ok: bool, msg: str)"""
                try:
                    import openpyxl as _xl
                    if f.name.endswith(".xlsx"):
                        wb = _xl.load_workbook(f, read_only=True)
                        f.seek(0)
                        sheets = [s.strip() for s in wb.sheetnames]
                        missing = [s for s in ["必有", "禁止"] if s not in sheets]
                        if missing:
                            return False, f"❌ **格式錯誤**：缺少工作表 `{'`, `'.join(missing)}`，請確認上傳的是「關鍵字」，而非音檔項目或題組"
                        return True, f"✅ 格式正確（含 `必有` / `禁止` 工作表）"
                    else:
                        return False, "❌ **格式錯誤**：關鍵字必須包含多個工作表（必有/禁止），請上傳 Excel (.xlsx) 格式，不支援 CSV。"
                except Exception as e:
                    return False, f"❌ 讀取失敗：{e}"
            
            # # ========== DEBUG 訊息 ==========
            # with st.expander("🐛 Debug 訊息（點擊展開）"):
            #     st.write("**Session State 檢查：**")
            #     st.write(f"- acceptance_id: {st.session_state.get('acceptance_id', '❌ 未設定')}")
            #     st.write(f"- is_clone: {st.session_state.get('is_clone', '❌ 未設定')}")
            #     st.write(f"- stt_done: {st.session_state.get('stt_done', '❌ 未設定')}")
            #     st.write(f"- log_keywords_path: {st.session_state.get('log_keywords_path', '❌ 未設定')}")
            #     st.write("\n**所有 Session State 變數：**")
            #     st.json({k: str(v)[:100] for k, v in st.session_state.items()})
            
            # ========== Step 3: 自訂關鍵字上傳 ==========
            # Clone 模式
            if st.session_state["is_clone"]:
                # 1. file_uploader 顯示上傳區
                # st.info("🟢 **關鍵字檔案 `keywords`**　｜　必要工作表：`必有`、`禁止`", icon="ℹ️")
                _kw_sample_path = os.path.join(BASE_DIR, "templates", "keywords_sample.xlsx")
                _kw_text = "請上傳關鍵字 keywords（僅支援 EXCEL，上限 200MB）"
                try:
                    with open(_kw_sample_path, "rb") as _f:
                        _b64 = base64.b64encode(_f.read()).decode()
                    _kw_icon = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{_b64}" download="keywords_sample.xlsx" title="下載範例檔" style="text-decoration: none;"><img src="https://cdn-icons-png.flaticon.com/512/732/732220.png" width="24" style="vertical-align: middle; margin-left: 10px;"></a>'
                except FileNotFoundError:
                    _kw_icon = '<span style="color:red; font-size:12px; margin-left:10px;">(未找到範例檔)</span>'
                st.markdown(f'<span style="font-size: 14px;">{_kw_text}{_kw_icon}</span>', unsafe_allow_html=True)
                uploaded_keywords = st.file_uploader(
                    _kw_text,
                    type=["xlsx"],
                    accept_multiple_files=True,  # 關鍵字只能上傳一個
                    label_visibility="collapsed",
                    key="keyword_uploader_clone")
                
                if uploaded_keywords:
                    for kw_file in uploaded_keywords:
                        _ok, _msg = _validate_keywords(kw_file)
                        if _ok:
                            st.success(f"{kw_file.name}: {_msg}")
                        else:
                            st.error(f"{kw_file.name}: {_msg}")

                # 2. 每次進入 Step3（Clone）都從 reuse_data 載入歷史關鍵字
                reuse_data = st.session_state.get("reuse_project_data", {})
                current_reuse_key = reuse_data.get("原始受理編號")

                # 3. 初始化 clone 模式的檔案物件
                # ③ 兩種情況要做初始化：
                #    A. 尚未初始化過（clone_kw_initialized旗標為 False）
                #    B. 上次初始化的對象不是這個專案（換了另一個「原始受理編號」）
                if (not st.session_state.get("clone_kw_initialized", False)) or \
                   (st.session_state.get("clone_kw_initialized_for") !=  current_reuse_key):
                    
                    st.session_state.clone_keyword_files = []
                    reuse_kw_raw = reuse_data.get("關鍵字路徑")
                    print(f"reuse_kw_raw:{reuse_kw_raw}") #可能是檔案路徑或 data URI / <a> 形式
                    st.session_state.clone_keyword_files = parse_reuse_keywords_to_fileobjs(reuse_kw_raw) #將「檔案路徑」或「data URI / <a>」轉成obj
                    print(f"st.session_state.clone_keyword_files:{st.session_state.clone_keyword_files}")
                    st.session_state.clone_kw_initialized = True #打上「已初始化」旗標，避免每次 rerun 都重載（這樣使用者按了❌刪除才會保留效果）
                    st.session_state.clone_kw_initialized_for = current_reuse_key #記錄這次初始化是針對哪一個「原始受理編號」

                    st.session_state.remove_clone_kw_indices = set() # 暫存使用者按❌想刪除的索引
                    st.session_state.run_remove_clone_kw = False # 旗標：下一輪 rerun 才真正執行刪除

                # 若上一輪按了任何 ❌，這輪先真正刪掉
                if st.session_state.run_remove_clone_kw:
                    to_remove = set(st.session_state.remove_clone_kw_indices)
                    # 重新過濾：只保留沒有被標記的檔案
                    st.session_state.clone_keyword_files = [ 
                        obj for i, obj in enumerate(st.session_state.clone_keyword_files)
                        if i not in to_remove
                    ]
                    st.session_state.remove_clone_kw_indices.clear()                                   # 清空待刪名單
                    st.session_state.run_remove_clone_kw = False                                      # 清旗標

                # (6) 顯示 clone 關鍵字清單（每一個都有 ❌ 可刪）
                st.markdown("🔁 載入歷史關鍵字：")
                if st.session_state.clone_keyword_files:
                    for idx, f in enumerate(st.session_state.clone_keyword_files):
                        c1, c2 = st.columns([8, 1])
                        with c1:
                            size_kb = len(f.getbuffer()) / 1024
                            st.markdown(f"📄 **{getattr(f, 'name', 'keywords.xlsx')}** ({size_kb:.1f} KB)")
                        with c2:
                            # 用唯一 key，避免與別的按鈕衝突
                            if st.button("❌", key=f"remove_clone_kw_{idx}"):
                                st.session_state.remove_clone_kw_indices.add(idx)
                                st.session_state.run_remove_clone_kw = True
                                st.rerun()  # 立即刷新 UI

                # (7) 確認按鈕：合併 cloned + uploaded，最後**必須只有 1 份**
                if st.button("確定使用關鍵字"):
                    # (7-1) 合併候選
                    merged = []
                    # 先放 cloned（可能 0~多個）
                    if st.session_state.clone_keyword_files:
                        merged.extend(st.session_state.clone_keyword_files)
                    # 再放上傳的（可能 0~多個）
                    if uploaded_keywords:
                        merged.extend(uploaded_keywords)
                    final_files = dedup_by_name(merged)
                    print(f"final_kw: {final_files}")

                    # (7-3) 驗證只能 1 份
                    if len(final_files) > 1:
                        st.warning("❗ 只能保留**一份**關鍵字檔案，請刪除或取消多餘的檔案後再確認。")
                    elif len(final_files) == 0:
                        st.warning("請上傳一份關鍵字檔案。")
                    else:
                        # (7-4) 落地唯一檔案
                        (file_obj,) = final_files  # 序列解包，同 final_files[0]
                        log_folder_path = init_project_log_folder(st.session_state.acceptance_id)
                        dst = os.path.join(log_folder_path, getattr(file_obj, "name", "keywords.xlsx"))
                        write_and_protect(dst, file_obj.read())

                        # (7-5) 更新 session（統一命名）
                        st.session_state.log_keywords_path = dst
                        st.session_state.uploaded_keywords = file_obj
                        st.session_state["keywords_dict"] = exceltodict(dst)

                        # (7-6) 寫 log
                        append_log(st.session_state.acceptance_id, f"上傳關鍵字：{dst}")

                        # (7-7) 寫 meta（若你已有工具函式）
                        meta_path = os.path.join(log_folder_path, "meta.json")
                        update_meta(meta_path, 'keyword_path', st.session_state.log_keywords_path)

                        # (7-8) 成功訊息
                        st.success(f"✅ 已上傳關鍵字：{os.path.basename(dst)}")

            # ========== Step 3 非 Clone ==========
            else:  # not is_clone
                # 1. 顯示上傳器
                # st.info("🟢 **關鍵字檔案 `keywords`**　｜　必要工作表：`必有`、`禁止`", icon="ℹ️")
                _kw_sample_path = os.path.join(BASE_DIR, "templates", "keywords_sample.xlsx")
                _kw_text = "請上傳關鍵字 keywords（僅支援 EXCEL，上限 200MB）"
                try:
                    with open(_kw_sample_path, "rb") as _f:
                        _b64 = base64.b64encode(_f.read()).decode()
                    _kw_icon = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{_b64}" download="keywords_sample.xlsx" title="下載範例檔" style="text-decoration: none;"><img src="https://cdn-icons-png.flaticon.com/512/732/732220.png" width="24" style="vertical-align: middle; margin-left: 10px;"></a>'
                except FileNotFoundError:
                    _kw_icon = '<span style="color:red; font-size:12px; margin-left:10px;">(未找到範例檔)</span>'
                st.markdown(f'<span style="font-size: 14px;">{_kw_text}{_kw_icon}</span>', unsafe_allow_html=True)
                uploaded_keywords = st.file_uploader(
                    _kw_text,
                    type=["xlsx"],
                    key="keyword_uploader_new",
                    label_visibility="collapsed",
                    accept_multiple_files=False)  # 明確宣告單檔
                
                if uploaded_keywords:
                    _ok, _msg = _validate_keywords(uploaded_keywords)
                    if _ok:
                        st.success(_msg)
                    else:
                        st.error(_msg)
                
                # 2. 確認使用
                if st.button("確定使用關鍵字檔"):
                    if uploaded_keywords is None:
                        st.warning("請上傳一份關鍵字檔案。")
                    else:
                        # 3. 寫入專案資料夾
                        log_folder_path = init_project_log_folder(st.session_state.acceptance_id)
                        filename = getattr(uploaded_keywords, "name", "keywords.xlsx")
                        log_keywords_path = os.path.join(log_folder_path, filename)
                        write_and_protect(log_keywords_path, uploaded_keywords.read())

                        # 4. 更新 session
                        st.session_state.uploaded_keywords = uploaded_keywords
                        st.session_state.log_keywords_path = log_keywords_path

                        # 5.解析成字典
                        st.session_state["keywords_dict"] = exceltodict(log_keywords_path) #解析關鍵字Excel成字典，存入session_state

                        # 5. 寫入 log
                        append_log(st.session_state.acceptance_id, f"上傳關鍵字：{log_keywords_path}")

                        # 6. 寫入 meta
                        meta_path = os.path.join(log_folder_path, "meta.json")
                        update_meta(meta_path, 'keyword_path', st.session_state.log_keywords_path)
                        
                        # 7. 成功訊息，並清除舊結果
                        st.success(f"✅ 已上傳關鍵字：{uploaded_keywords.name}")
                        for k in ["show_results", "log_result_path"]:
                            st.session_state.pop(k, None)        

            # ========== Step 4: 比對結果彙整 ==========
            st.markdown('----------------------------------')
            st.markdown('<p style="font-size: 36px; font-weight: bold;">彙整比對結果</p>', unsafe_allow_html=True)

            
            # 初始化查詢觸發 flag
            if "df_results_query_triggered" not in st.session_state:
                st.session_state.df_results_query_triggered = False
            if "last_df_filters" not in st.session_state:
                st.session_state.last_df_filters = {
                    "chars": [],
                    "sources": [],
                    "keywords": [],
                    "matches": [],
                    "paths": []
                }
            if "show_filter" not in st.session_state:
                st.session_state.show_filter = False
            if "df_results_filter_confirmed" not in st.session_state:
                st.session_state.df_results_filter_confirmed = False

            # 檢查是否已完成前置步驟
            transcript_paths = st.session_state.get("transcript_paths", [])
            keywords_path = st.session_state.get("log_keywords_path", None)
            
            # 顯示前置步驟完成狀態
            if not transcript_paths:
                st.warning("⚠️ 請先在「頁籤一：產出逐字稿」完成 Step 2 語音轉文字（STT）")
            
            if not keywords_path:
                st.warning("⚠️ 請先完成 Step 3 上傳關鍵字檔案")
            
            # 設定 transcript_path 和 keyword_path
            print(f"transcript_paths:{transcript_paths}")
            print(f"keywords_path:{keywords_path}")

            if "kw_btn_disabled" not in st.session_state:
                st.session_state.kw_btn_disabled = False
            # 切換 tab 再回來：若上次比對已完成，自動解鎖讓使用者可以重新執行
            elif st.session_state.get("show_results", False):
                st.session_state.kw_btn_disabled = False

            def disable_kw_btn():
                st.session_state.kw_btn_disabled = True

            if keywords_path and transcript_paths and st.button("執行比對", disabled=st.session_state.kw_btn_disabled, on_click=disable_kw_btn):
                
                with st.spinner("正在比對中..."):
                    df_results = run_callcompliance(transcript_paths, keywords_path, debug=True)
                print(df_results)
                #轉為字串並處理空值欄位
                df_results = df_results.fillna("").astype(str)
                st.session_state["df_results"] = df_results  # 存到 session_state
                st.session_state.df_results_query_triggered = False  # 預設不立刻顯示查詢結果
                
                # 匯出檔名 & 本機儲存路徑
                export_filename = f"比對結果_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
                log_folder_path = init_project_log_folder(st.session_state.acceptance_id)
                export_path = os.path.join(log_folder_path, export_filename)
                
                # 顯示提示訊息
                st.session_state["success_msg"] = f"比對完成，已儲存至比對結果路徑"
                st.session_state["df_results_info"] = pd.DataFrame([{
                    "結果路徑": export_path,
                    "比對時間": datetime.now().strftime("%Y/%m/%d %H:%M:%S")
                }])
                st.session_state["show_results"] = True  # 額外記錄「比對完成」

                # 直接存入 project_logs 資料夾，不寫入 result
                excel_bytes = io.BytesIO() #建立記憶體中暫存區
                df_results.to_excel(excel_bytes, index=False) #寫入暫存區
                write_and_protect(export_path, excel_bytes.getvalue()) #取出暫存區中Excel 
                append_log(st.session_state.acceptance_id, f"比對完成：{export_path}")
                st.session_state.log_result_path = export_path           

                # 路徑標準化（全都 list）
                norm_transcript_paths = _norm_paths(st.session_state.transcript_paths)
                norm_log_keywords_path = _norm_paths(st.session_state.log_keywords_path)
                norm_log_result_path = _norm_paths(st.session_state.log_result_path)
                
                # 加入歷史紀錄
                st.session_state.history = st.session_state.get("history", [])
                st.session_state.history.append({
                    "受理編號": st.session_state.acceptance_id,
                    "員工編號":st.session_state.employee_id,
                    "逐字稿路徑清單": norm_transcript_paths, #list 檔案路徑清單
                    "關鍵字路徑": norm_log_keywords_path, #str 檔案路徑
                    "比對結果路徑": norm_log_result_path #str 檔案路徑
                })
                
                st.session_state.kw_btn_disabled = False
                st.rerun()
                
            # ➤ 「永遠顯示」比對成功訊息（只要有 show_results=True）
            if st.session_state.get("show_results", False):
                st.success(st.session_state.get("success_msg", ""))
            
            if st.session_state.get("show_results", False):
                st.markdown('<p style="font-size: 24px; font-weight: bold;">比對結果路徑</p>', unsafe_allow_html=True)
                st.table(st.session_state.get("df_results_info", pd.DataFrame()))
            
                # ➤ 立即下載 Excel（直接顯示下載鈕）
                if st.session_state.get("log_result_path") and os.path.exists(st.session_state["log_result_path"]):
                    with open(st.session_state["log_result_path"], "rb") as f:
                        st.download_button(
                            label="⬇️⬇️⬇️ 立即下載比對結果 Excel",
                            data=f,
                            file_name=os.path.basename(st.session_state["log_result_path"]),
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )


                # ✅ 顯示一個新按鈕，用來顯示篩選器
                if st.button("使用篩選器"):
                    st.session_state.show_filter = True

                #只有執行過篩選後，才進行下列動作
                if st.session_state.get("show_filter", False) and "df_results" in st.session_state:
                    df_results = st.session_state["df_results"]
                    st.markdown('<p style="font-size: 24px; margin-top: 30px; font-weight: bold;">篩選器</p>', unsafe_allow_html=True)

                    # 建立4個多選選單
                    col1, col2, col3, col4 = st.columns(4)
                    with col1:
                        selected_chars = st.multiselect("角色", df_results["角色"].unique())
                    with col2:
                        selected_sources = st.multiselect("來源", df_results["來源"].unique())
                    # 動態關鍵字選單內容
                    keywords_dict = st.session_state.get("keywords_dict", {})
                    keyword_options = set()  # 用 set 去重

                    if selected_sources and selected_chars:
                        # 來源 + 角色都有選時，交叉找出所有符合組合
                        for sheet in selected_sources:
                            if sheet in keywords_dict:
                                for topic in selected_chars:
                                    if topic in keywords_dict[sheet]:
                                        keyword_options.update(keywords_dict[sheet][topic])
                    elif selected_chars:
                        # 只有選角色（來源未選），就列出所有來源下該角色的關鍵字
                        for sheet, sheet_data in keywords_dict.items():
                            for topic in selected_chars:
                                if topic in sheet_data:
                                    keyword_options.update(sheet_data[topic])
                    elif selected_sources:
                        # 只有選來源（角色未選），就列出該來源下所有角色的關鍵字
                        for sheet in selected_sources:
                            if sheet in keywords_dict:
                                for topic_keywords in keywords_dict[sheet].values():
                                    keyword_options.update(topic_keywords)
                    else:
                        # 角色 + 來源都沒選，列出所有關鍵字
                        for sheet_data in keywords_dict.values():
                            for topic_keywords in sheet_data.values():
                                keyword_options.update(topic_keywords)

                    keyword_options = sorted(keyword_options)  # 排序
                    with col3:
                        selected_keywords = st.multiselect("關鍵字", options=keyword_options)
                    with col4:
                        selected_matches = st.multiselect("是否比對到", df_results["是否比對到"].unique())

                    all_paths = df_results["逐字稿名稱"].unique()
                    selected_set = set(selected_chars)

                    # 根據單位設定對應方向
                    if st.session_state.unit_code == "1" or "3":  # 服中、直效預設
                        ag_direction = "left"
                        cus_direction = "right"
                    elif st.session_state.unit_code == "2":  # 智客
                        ag_direction = "right"
                        cus_direction = "left"
                    # else:
                    #     ag_direction = None  # 其他單位顯示全部
                    #     cus_direction = None

                    # 擴大角色集
                    ag_roles = {"業務", "客服", "理專"}

                    # 根據角色與單位決定 path_options
                    print(f'{st.session_state.unit_code}單位')
                    if ag_direction and selected_set.issubset(ag_roles) and selected_set:
                        print(f'在ag_roles中:{selected_set.issubset(ag_roles)}, 選角色:{selected_set}，選{ag_direction}')
                        path_options = [p for p in all_paths if ag_direction in p]
                    elif selected_set == {"客戶"}:
                        print(f'選角色:{selected_set}，選{cus_direction}')
                        path_options = [p for p in all_paths if cus_direction in p]
                    else:
                        print(f'選角色:{selected_set}，全選')
                        path_options = list(all_paths)
                    
                    selected_paths = path_options  # 永遠使用全部，不顯示選單

                    # 在 multiselect 完成後檢查欄位是否變更
                    current_df_filters = {
                        "chars": selected_chars,
                        "sources": selected_sources,
                        "keywords": selected_keywords,
                        "matches": selected_matches,
                        "paths": selected_paths
                    }
                    if current_df_filters != st.session_state.last_df_filters:
                        st.session_state.df_results_query_triggered  = False  # 一變就收起
                        st.session_state.last_df_filters = current_df_filters

                    # 確認查詢按鈕
                    if st.button("確認查詢"):
                        st.session_state.df_results_query_triggered = True
                        st.session_state.df_results_filter_confirmed = False  # 每次重新查詢 → 重置確認狀態

                    # 若按下確認查詢才顯示表格
                    if st.session_state.df_results_query_triggered:
                        #根據選擇的條件進行篩選
                        filtered_df = df_results.copy()
                        if selected_chars:
                            filtered_df = filtered_df[filtered_df["角色"].isin(selected_chars)]
                        if selected_sources:
                            filtered_df = filtered_df[filtered_df["來源"].isin(selected_sources)]
                        if selected_keywords:
                            filtered_df = filtered_df[filtered_df["關鍵字"].isin(selected_keywords)]
                        if selected_matches:
                            filtered_df = filtered_df[filtered_df["是否比對到"].isin(selected_matches)]
                        if selected_paths:
                            filtered_df = filtered_df[filtered_df["逐字稿名稱"].isin(selected_paths)]

                        # 顯示結果：改用 st.dataframe，自動調整欄寬

                        # 建立顯示 DataFrame
                        display_df = filtered_df[["角色", "來源", "關鍵字", "是否比對到", "逐字稿時間"]].copy()
                        display_df.insert(0, "檔案名稱",
                            filtered_df["檔案"].apply(lambda x: os.path.basename(os.path.normpath(str(x)))))
                        display_df["檔案下載"] = filtered_df["檔案"].apply(
                            lambda x: local_path_to_http_url(os.path.normpath(str(x))))

                        st.dataframe(
                            display_df,
                            use_container_width=True,
                            hide_index=True,
                            column_config={
                                "檔案下載": st.column_config.LinkColumn(
                                    "檔案下載",
                                    display_text=r"([^/]+)$"
                                )
                            }
                        )


                        if st.button("確認篩選條件"):
                            # 統一 timestamp
                            now_str = datetime.now().strftime("%Y%m%d_%H%M%S")
                            export_filename_filtered = f"篩選結果_{now_str}.xlsx"
                            
                            log_folder_path_filtered = init_project_log_folder(st.session_state.acceptance_id)
                            export_path_filtered = os.path.join(log_folder_path_filtered, export_filename_filtered)

                            # 建立記憶體下載版
                            excel_bytes_filtered = io.BytesIO()
                            filtered_df.to_excel(excel_bytes_filtered, index=False)
                            excel_bytes_filtered.seek(0)

                            # ✅ 直接儲存至專案資料夾 (project_logs)
                            write_and_protect(export_path_filtered, excel_bytes_filtered.getvalue())
                            
                            # ✅ 存到 session_state，避免 rerun 消失
                            st.session_state["excel_bytes_filtered"] = excel_bytes_filtered.getvalue()
                            st.session_state["export_filename_filtered"] = export_filename_filtered
                            st.session_state["export_path_filtered"] = export_path_filtered

                            # 顯示儲存結果與路徑
                            st.session_state["filtered_df_results_info"] = pd.DataFrame([{
                                "篩選路徑": export_path_filtered,
                                "篩選時間": datetime.now().strftime("%Y/%m/%d %H:%M:%S")
                            }])
                            st.session_state.df_results_filter_confirmed = True

                            append_log(st.session_state.acceptance_id, f"篩選完成：{export_path_filtered}")             
                            protect_file(os.path.join(log_folder_path_filtered, "操作紀錄.log"))

                # 4) 篩選完成後 → 永遠顯示下載與路徑
                if st.session_state.get("df_results_filter_confirmed", False):

                    st.markdown('<p style="font-size: 24px; font-weight: bold;">篩選結果路徑</p>', unsafe_allow_html=True)
                    st.table(st.session_state.get("filtered_df_results_info", pd.DataFrame()))

                    # 顯示下載按鈕（用同一個檔名）
                    st.success(f"篩選完成，已儲存至篩選結果路徑")
                    st.download_button(
                        label="⬇️⬇️⬇️ 立即下載篩選結果 Excel",
                        data=st.session_state["excel_bytes_filtered"],
                        file_name=st.session_state.get("export_filename_filtered", "篩選結果.xlsx"),
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )


        # ==================== 頁籤三：AI 質檢 ====================
        with tab3:
            st.markdown('<p style="font-size: 36px; font-weight: bold;">設定模型</p>', unsafe_allow_html=True)
            
            # # ========== DEBUG 訊息 ==========
            # with st.expander("🐛 Debug 訊息（點擊展開）"):
            #     st.write("**Session State 檢查：**")
            #     st.write(f"- acceptance_id: {st.session_state.get('acceptance_id', '❌ 未設定')}")
            #     st.write(f"- qa_model_choice: {st.session_state.get('qa_model_choice', '❌ 未設定')}")
            #     st.write(f"- qa_question_path: {st.session_state.get('qa_question_path', '❌ 未設定')}")
            #     st.write(f"- qa_audioitem_path: {st.session_state.get('qa_audioitem_path', '❌ 未設定')}")
            #     st.write(f"- qa_done: {st.session_state.get('qa_done', '❌ 未設定')}")
            #     st.write("\n**所有 Session State 變數：**")
            #     st.json({k: str(v)[:100] for k, v in st.session_state.items()})
            
            # 初始化 session state
            if "qa_model_choice" not in st.session_state:
                st.session_state.qa_model_choice = "Gemini"
            if "qa_gemma_api_url" not in st.session_state:
                st.session_state.qa_gemma_api_url = "http://10.61.16.101:8000/v1/chat/completions"
            if "qa_gemma_model_name" not in st.session_state:
                st.session_state.qa_gemma_model_name = "/mnt/model/gemma-3-27b-it-qat-compressed-tensors"
            if "qa_temperature" not in st.session_state:
                st.session_state.qa_temperature = 0.0
            if "qa_done" not in st.session_state:
                st.session_state.qa_done = False
            if "qa_ollama_api_url" not in st.session_state:
                st.session_state.qa_ollama_api_url = "http://10.61.16.31:11434/api/generate"
            if "qa_ollama_model_name" not in st.session_state:
                st.session_state.qa_ollama_model_name = "gpt-oss:20b"

            model_choice = st.radio("選擇 LLM 模型", ["Gemini", "AzureGPT"], key="qa_model_choice")

            # 切換模型時清空上一次的質檢結果
            if st.session_state.get("qa_last_model_choice") != model_choice:
                if st.session_state.get("qa_last_model_choice") is not None:  # 非首次進入
                    for _k in [
                        "qa_done", "qa_results",
                        "qa_show_filter",
                        "qa_confirm_query", "qa_confirmed_status", "qa_confirmed_audio_ids",
                        "qa_filter_confirmed", "qa_filter_excel_bytes", "qa_filter_export_filename",
                        "qa_filter_export_path", "qa_filter_export_info",
                        "qa_current_page", "qa_last_filter", "qa_report_path", "qa_llm_label",
                    ]:
                        st.session_state.pop(_k, None)
                    st.session_state.qa_btn_disabled = False  # ← 切換模型時解鎖按鈕
                    st.info(f"🔄 已切換至 {model_choice}，請重新執行質檢")
                st.session_state["qa_last_model_choice"] = model_choice

            # if model_choice == "Gemini":
            #     st.info("使用 Google Gemini API（gemini-2.5-flash）")
            #     st.caption("將使用預設 API Key 進行質檢")
            # else:
            #     api_url = st.text_input(
            #         "Gemma API URL", 
            #         value=st.session_state.qa_gemma_api_url,
            #         key="qa_gemma_api_url"
            #     )
            #     model_name = st.text_input(
            #         "模型名稱", 
            #         value=st.session_state.qa_gemma_model_name,
            #         key="qa_gemma_model_name"
            #     )
            #     temperature = st.slider(
            #         "Temperature", 
            #         0.0, 1.0, 
            #         st.session_state.qa_temperature,
            #         key="qa_temperature"
            #     )
            
            # ── Clone 模式：自動帶入原始專案的題組設定 ──
            if st.session_state.get("is_clone", False):
                reuse_data = st.session_state.get("reuse_project_data") or {}
                src_acceptance_id = reuse_data.get("原始受理編號") or reuse_data.get("acceptance_id", "")
                
                # 初始化：如果還沒載入過 Clone 的檔案，就帶過來
                if src_acceptance_id and not st.session_state.get("qa_clone_initialized"):
                    src_project_folder = init_project_log_folder(src_acceptance_id)
                    src_qa_config_dir  = os.path.join(src_project_folder, "qa_config")
                    
                    # 嘗試從來源專案的 meta.json 讀取真正儲存的檔名
                    src_meta_path = os.path.join(src_project_folder, "meta.json")
                    src_meta = {}
                    if os.path.exists(src_meta_path):
                        try:
                            with open(src_meta_path, "r", encoding="utf-8") as f:
                                src_meta = json.load(f)
                        except:
                            pass
                    
                    # 優先使用 meta 記錄的原始路徑（若不存在則 fallback 回預設檔名）
                    src_audioitem = src_meta.get("qa_audioitem_path")
                    if not src_audioitem or not os.path.exists(src_audioitem):
                        src_audioitem = os.path.join(src_qa_config_dir, "audioitem.xlsx")
                        
                    src_questionset = src_meta.get("qa_question_path")
                    if not src_questionset or not os.path.exists(src_questionset):
                        src_questionset = os.path.join(src_qa_config_dir, "questionset.xlsx")
                    
                    if os.path.exists(src_audioitem) and os.path.exists(src_questionset):
                        # 複製到當前專案，保持原本的檔名
                        project_folder  = init_project_log_folder(st.session_state.acceptance_id)
                        dst_qa_config   = os.path.join(project_folder, "qa_config")
                        os.makedirs(dst_qa_config, exist_ok=True)
                        
                        dst_audioitem   = os.path.join(dst_qa_config, os.path.basename(src_audioitem))
                        dst_questionset = os.path.join(dst_qa_config, os.path.basename(src_questionset))
                        
                        shutil.copy2(src_audioitem,  dst_audioitem)
                        shutil.copy2(src_questionset, dst_questionset)
                        
                        st.session_state.qa_audioitem_path_temp = dst_audioitem
                        st.session_state.qa_question_path_temp  = dst_questionset
                        st.session_state.qa_clone_initialized = True

            # ========== Step 6: 題組設定 ==========
            st.markdown('----------------------------------')
            st.markdown('<p style="font-size: 36px; font-weight: bold;">上傳音檔項目及題組</p>', unsafe_allow_html=True)
            
            # st.write("請  上傳 QA 設定檔案（支援 EXCEL, CSV；上限200MB）")

            # ── 防呆輔助函式 ──
            def _validate_audioitem(f):
                """回傳 (ok: bool, msg: str)"""
                try:
                    import pandas as pd
                    import os
                    # 為了檢查內容，必須讀取完整檔案，不能限制 nrows=1
                    df = pd.read_excel(f) if f.name.endswith(".xlsx") else pd.read_csv(f)
                    f.seek(0)
                    cols = [c.strip().lower() for c in df.columns]
                    missing = [c for c in ["audio_id", "item"] if c not in cols]
                    if missing:
                        return False, f"❌ **格式錯誤**：缺少必要欄位 `{'`, `'.join(missing)}`，請確認上傳的是「音檔項目」，而非題組或關鍵字"
                    
                    # 檢查是否有空值 (排除完全空白的列)
                    # 先刪除全部都是 NaN 的列，再檢查特定欄位是否有空值
                    df_clean = df.dropna(how='all', subset=['audio_id', 'item'])
                    if df_clean[['audio_id', 'item']].isna().any().any() or (df_clean[['audio_id', 'item']] == "").any().any():
                         return False, "❌ **內容錯誤**：`audio_id` 或 `item` 欄位中存在空白值，請確保每一筆音檔對應的作業項目都已完整填寫。"
                    
                    # 檢查上傳的音檔是否都有出現在 audio_id 欄位
                    uploaded_audios = st.session_state.get("audio_file_names", [])
                    if uploaded_audios:
                        # 將 audio_id 欄位的所有值轉為字串去空白，並準備一個小寫版供比對
                        # Excel裡面的 audio_id 通常「沒有」副檔名
                        audio_ids_str = df["audio_id"].astype(str).str.strip().tolist()
                        audio_ids_lower = [a.lower() for a in audio_ids_str]
                        
                        missing_audios = []
                        for audio_name in uploaded_audios:
                            # 上傳的檔案名稱去副檔名（例如 01.ABA000..._12-49.mp3 -> 01.ABA000..._12-49）
                            audio_name_no_ext = os.path.splitext(audio_name)[0]
                            audio_name_no_ext_lower = audio_name_no_ext.lower()
                            
                            # 檢查「去副檔名」後的主檔名，是否存在於 audio_id 欄位中 (不分大小寫比對)
                            if audio_name_no_ext_lower not in audio_ids_lower:
                                missing_audios.append(audio_name)
                        
                        if missing_audios:
                            # 錯誤訊息中，列出找不到的「完整音檔名」，讓使用者知道是哪個檔案被判定為遺漏
                            return False, f"❌ **內容不符**：上傳的音檔 `{'`, `'.join(missing_audios)}` 未出現在 `audio_id` 欄位中，請檢查清單。"

                    return True, f"✅ 格式正確（含 `audio_id` / `item` 欄，且音檔有對應）"
                except Exception as e:
                    return False, f"❌ 讀取失敗：{e}"

            def _validate_questionset(f):
                """回傳 (ok: bool, msg: str)"""
                try:
                    import openpyxl as _xl
                    import pandas as pd
                    
                    if not f.name.endswith(".xlsx"):
                        return False, "❌ **格式錯誤**：題組必須包含多個工作表（fixed/flexible），請上傳 Excel (.xlsx) 格式，不支援 CSV。"
                        
                    wb = _xl.load_workbook(f, read_only=True)
                    f.seek(0)
                    sheets = [s.strip().lower() for s in wb.sheetnames]
                    missing = [s for s in ["fixed", "flexible"] if s not in sheets]
                    if missing:
                        return False, f"❌ **格式錯誤**：缺少工作表 `{'`, `'.join(missing)}`，請確認上傳的是「題組」，而非音檔項目或關鍵字"
                        
                    # 1. 檢查 fixed 頁籤
                    try:
                        df_fixed = pd.read_excel(f, sheet_name="fixed")
                        # 必填欄位檢查
                        req_cols_fixed = ["question_category", "question_text"]
                        missing_f = [c for c in req_cols_fixed if c not in df_fixed.columns]
                        if missing_f:
                            return False, f"❌ **格式錯誤**：`fixed` 工作表缺少必要欄位 `{'`, `'.join(missing_f)}`"
                            
                        # 檢查 fixed 是否有空值
                        df_f_clean = df_fixed.dropna(how='all', subset=req_cols_fixed)
                        if df_f_clean[req_cols_fixed].isna().any().any() or (df_f_clean[req_cols_fixed] == "").any().any():
                            return False, "❌ **內容錯誤**：題組 `fixed` 工作表中存在空白格子，請確保每一題的分類與內容都已完整填寫。"
                    except Exception as e:
                        return False, f"❌ 無法讀取題組的 fixed 工作表：{e}"

                    # 第二階段：檢查 flexible 頁籤裡面的 item，是否涵蓋了 audio_item 裡面的 item
                    try:
                        f.seek(0)
                        df_flex = pd.read_excel(f, sheet_name="flexible")
                        
                        req_cols_flex = ["item", "question_category", "question_text"]
                        missing_fl = [c for c in req_cols_flex if c not in df_flex.columns]
                        if missing_fl:
                            return False, f"❌ **格式錯誤**：題組的 `flexible` 工作表中缺少必要欄位 `{'`, `'.join(missing_fl)}`"

                        # 檢查 flexible 是否有空值
                        df_fl_clean = df_flex.dropna(how='all', subset=req_cols_flex)
                        if df_fl_clean[req_cols_flex].isna().any().any() or (df_fl_clean[req_cols_flex] == "").any().any():
                            return False, "❌ **內容錯誤**：題組 `flexible` 工作表中存在空白格子，請確保每一題的作業項目、分類與內容都已完整填寫。"

                        qs_items = set(df_fl_clean["item"].astype(str).str.strip())
                    except Exception as e:
                        return False, f"❌ 無法讀取題組的 flexible 工作表：{e}"
                        
                    # 2. 嘗試取得目前上傳或歷史的 audio_item 的資料
                    audio_item_file = st.session_state.get("qa_audioitem_file_uploader")
                    audio_item_path = st.session_state.get("qa_audioitem_path_temp")
                    
                    df_audio = None
                    try:
                        if audio_item_file:
                            audio_item_file.seek(0)
                            df_audio = pd.read_excel(audio_item_file) if audio_item_file.name.endswith(".xlsx") else pd.read_csv(audio_item_file)
                        elif audio_item_path and os.path.exists(audio_item_path) and not st.session_state.get("qa_remove_audioitem"):
                            df_audio = pd.read_excel(audio_item_path) if audio_item_path.endswith(".xlsx") else pd.read_csv(audio_item_path)
                    except Exception as e:
                        pass # 若無法讀取音檔項目，則跳過此項交叉比對
                        
                    # 3. 如果成功取得音檔項目，進行交集比對
                    if df_audio is not None and "item" in df_audio.columns:
                        audio_items = set(df_audio["item"].dropna().astype(str).str.strip())
                        missing_items = audio_items - qs_items
                        if missing_items:
                            return False, f"❌ **項目短少**：音檔項目 (`audio_item`) 中出現了 `{'`, `'.join(missing_items)}`，但題組檔案的 `flexible` 工作表中沒有定義這些項目的題目！請修正。"

                    return True, f"✅ 格式正確（含 `fixed` / `flexible` 且項目吻合）"
                    
                except Exception as e:
                    return False, f"❌ 讀取失敗：{e}"

            # 1. 顯示上傳器（加上視覺區塊與即時驗證）

            # 準備下載用的 Base64 檔案連結產生函式，以及 Icon 圖片
            import base64
            def get_download_link(file_path, file_name, link_text=""):
                try:
                    with open(file_path, "rb") as f:
                        b64 = base64.b64encode(f.read()).decode()
                        # 使用一個簡單但好看的 Excel SVG Icon 或者是圖片網址
                        icon_url = "https://cdn-icons-png.flaticon.com/512/732/732220.png"
                        return f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{file_name}" title="下載範例檔" style="text-decoration: none;"><img src="{icon_url}" width="24" style="vertical-align: middle; margin-left: 10px;"> {link_text}</a>'
                except FileNotFoundError:
                    return '<span style="color:red; font-size:12px; margin-left:10px;">(未找到範例檔)</span>'

            # ── 藍色區塊：audio_item ──
            audio_sample_path = os.path.join(BASE_DIR, "templates", "audioitem_sample.xlsx")
            audio_text = f"請上傳音檔項目 audioitem（支援 EXCEL, CSV，上限 200MB）"
            st.markdown(f'<span style="font-size: 14px;">{audio_text}{get_download_link(audio_sample_path, "audioitem_sample.xlsx")}</span>', unsafe_allow_html=True)
            
            audioitem_file = st.file_uploader(
                audio_text,
                label_visibility="collapsed",
                type=["xlsx", "csv"],
                key="qa_audioitem_file_uploader"
            )
            if audioitem_file:
                _ok, _msg = _validate_audioitem(audioitem_file)
                if _ok:
                    st.success(_msg)
                else:
                    st.error(_msg)

            # ── 橘色區塊：questionset ──
            qs_sample_path = os.path.join(BASE_DIR, "templates", "questionset_sample.xlsx")
            qs_text = f"請上傳題組 questionset（僅支援 EXCEL，上限 200MB）"
            st.markdown(f'<span style="font-size: 14px;">{qs_text}{get_download_link(qs_sample_path, "questionset_sample.xlsx")}</span>', unsafe_allow_html=True)
            
            question_file = st.file_uploader(
                qs_text,
                label_visibility="collapsed",
                type=["xlsx"],
                key="qa_question_file_uploader"
            )
            if question_file:
                _ok, _msg = _validate_questionset(question_file)
                if _ok:
                    st.success(_msg)
                else:
                    st.error(_msg)
                    # st.caption("💡 請確認上傳的是「題組」，而非音檔項目或關鍵字檔")

            # 2. 顯示歷史/Clone帶入的檔案（附刪除按鈕）
            if st.session_state.get("is_clone") and (st.session_state.get("qa_audioitem_path_temp") or st.session_state.get("qa_question_path_temp")):
                st.markdown("🔁 載入歷史設定：")
                
                # 刪除機制
                if st.session_state.get("qa_remove_audioitem"):
                    st.session_state.qa_audioitem_path_temp = None
                    st.session_state.qa_remove_audioitem = False
                if st.session_state.get("qa_remove_question"):
                    st.session_state.qa_question_path_temp = None
                    st.session_state.qa_remove_question = False

                if st.session_state.get("qa_audioitem_path_temp"):
                    c1, c2 = st.columns([8, 1])
                    with c1:
                        st.markdown(f"📄 **{os.path.basename(st.session_state.qa_audioitem_path_temp)}** (音檔項目)")
                    with c2:
                        if st.button("❌", key="btn_remove_qa_audioitem"):
                            st.session_state.qa_remove_audioitem = True
                            st.rerun()
                            
                if st.session_state.get("qa_question_path_temp"):
                    c1, c2 = st.columns([8, 1])
                    with c1:
                        st.markdown(f"📄 **{os.path.basename(st.session_state.qa_question_path_temp)}** (題組)")
                    with c2:
                        if st.button("❌", key="btn_remove_qa_question"):
                            st.session_state.qa_remove_question = True
                            st.rerun()

            # 3. 確定使用按鈕
            if st.button("確定使用設定檔", key="btn_confirm_qa_settings"):
                if not st.session_state.get("acceptance_id"):
                    st.error("❌ 請先建立專案（Step 0）")
                else:
                    project_folder = init_project_log_folder(st.session_state.acceptance_id)
                    qa_config_dir = os.path.join(project_folder, "qa_config")
                    os.makedirs(qa_config_dir, exist_ok=True)

                    # 處理 audioitem
                    final_audioitem_path = None
                    if audioitem_file:
                        final_audioitem_path = os.path.join(qa_config_dir, audioitem_file.name)
                        with open(final_audioitem_path, "wb") as f:
                            f.write(audioitem_file.getbuffer())
                    elif st.session_state.get("qa_audioitem_path_temp"):
                        final_audioitem_path = st.session_state.qa_audioitem_path_temp

                    # 處理 questionset
                    final_question_path = None
                    if question_file:
                        final_question_path = os.path.join(qa_config_dir, question_file.name)
                        with open(final_question_path, "wb") as f:
                            f.write(question_file.getbuffer())
                    elif st.session_state.get("qa_question_path_temp"):
                        final_question_path = st.session_state.qa_question_path_temp

                    # 驗證
                    if not final_audioitem_path or not final_question_path:
                        # 反映當下空白狀態：清除 qa_config 與 session_state
                        if os.path.exists(qa_config_dir):
                            for _f in os.listdir(qa_config_dir):
                                _fp = os.path.join(qa_config_dir, _f)
                                if os.path.isfile(_fp):
                                    try:
                                        os.remove(_fp)
                                    except Exception as e:
                                        print(f"清除 qa_config 失敗 {_f}: {e}")
                        st.session_state.pop("qa_audioitem_path", None)
                        st.session_state.pop("qa_question_path", None)
                        st.session_state.pop("qa_settings_confirmed", None)
                        st.warning("❗ 請確保已上傳/帶入「音檔項目」與「題組」兩份檔案。")
                    else:
                        st.session_state.qa_audioitem_path = final_audioitem_path
                        st.session_state.qa_question_path = final_question_path
                        
                        # 🧹 同步清理資料夾：如果 qa_config_dir 內有未被使用的舊檔案（如被新上傳覆蓋的舊 Clone 檔），則刪除之
                        if os.path.exists(qa_config_dir):
                            for f in os.listdir(qa_config_dir):
                                f_path = os.path.join(qa_config_dir, f)
                                if os.path.isfile(f_path) and f_path not in [final_audioitem_path, final_question_path]:
                                    try:
                                        os.remove(f_path)
                                    except Exception as e:
                                        print(f"無法刪除未使用的設定檔 {f}: {e}")
                        
                        append_log(st.session_state.acceptance_id, f"設定題組檔案：{final_question_path}")
                        append_log(st.session_state.acceptance_id, f"設定 audio_item：{final_audioitem_path}")

                        st.session_state.qa_settings_confirmed = True
                        st.session_state.qa_btn_disabled = False
                        st.success(f"""✅ 已套用設定：
- 音檔項目：{os.path.basename(final_audioitem_path)}
- 題組：{os.path.basename(final_question_path)}""")
            
            # ========== Step 7: 執行質檢 ==========
            st.markdown('----------------------------------')
            st.markdown('<p style="font-size: 36px; font-weight: bold;">📋 彙整質檢結果</p>', unsafe_allow_html=True)
            
            # 自動找到專案資料夾中的所有 JSON 檔案
            if st.session_state.get("acceptance_id"):
                project_folder = init_project_log_folder(st.session_state.acceptance_id)
                transcript_folder = os.path.join(project_folder, "逐字稿")
                
                # 尋找所有 JSON 檔案（在逐字稿子資料夾中搜尋）
                import pathlib
                json_files = []
                if os.path.exists(transcript_folder):
                    # 使用 rglob 遞迴搜尋所有符合條件的 JSON 檔案
                    json_files = list(pathlib.Path(transcript_folder).rglob(f"*_transcript_openaiwhisper_azureapi.json"))
                    # 只保留主音檔，排除切音軌檔案（_right / _left）
                    json_files = [f for f in json_files if "_right_" not in f.stem and "_left_" not in f.stem]
                

                # # Debug: 顯示搜尋路徑
                # with st.expander("🔍 搜尋路徑資訊（Debug）"):
                #     st.write(f"**專案資料夾：** `{project_folder}`")
                #     st.write(f"**逐字稿資料夾：** `{transcript_folder}`")
                #     st.write(f"**資料夾是否存在：** {os.path.exists(transcript_folder)}")
                #     st.write(f"**搜尋模式：** `*_transcript_openaiwhisper_large.json`")
                #     if os.path.exists(transcript_folder):
                #         all_files = list(pathlib.Path(transcript_folder).rglob("*.json"))
                #         st.write(f"**找到的所有 JSON 檔案：** {len(all_files)} 個")
                #         for f in all_files:
                #             st.caption(f"  • {f}")
                #     else:
                #         st.warning(f"⚠️ 逐字稿資料夾不存在：{transcript_folder}")
                #         st.warning("⚠️ 請先在「頁籤一：產出逐字稿」完成 Step 2 語音轉文字（STT）")
                
                if json_files:
                    st.info(f"📁 找到 {len(json_files)} 個主音檔逐字稿 JSON 檔案")
                    
                    # 顯示檔案列表（可展開）
                    with st.expander("查看檔案列表"):
                        for f in json_files:
                            st.caption(f"• {f.name}")
                else:
                    st.warning("⚠️ 專案資料夾中沒有找到逐字稿 JSON 檔案")
                    st.warning("⚠️ 請先在「頁籤一：產出逐字稿」完成 Step 2 語音轉文字（STT）")
            
            # ====== 預估等待時間計算與顯示 ======
            if st.session_state.get("qa_question_path") and st.session_state.get("qa_audioitem_path") and json_files:
                try:
                    import pandas as pd
                    # 1. 取得題組與題數
                    df_fixed = pd.read_excel(st.session_state.qa_question_path, sheet_name="fixed")
                    df_flex = pd.read_excel(st.session_state.qa_question_path, sheet_name="flexible")
                    num_fixed = len(df_fixed)
                    flex_counts = df_flex.groupby("item").size().to_dict()
                    
                    # 2. 取得音檔清單映射
                    df_audio = pd.read_excel(st.session_state.qa_audioitem_path)
                    
                    # 3. 計算總題數
                    total_q = 0
                    file_q_counts = {}  # 記錄每一個 json 檔被分配到幾題，用來 debug 和寫 log
                    
                    # 預先建立乾淨的 audio_id -> item 對照表，加速比對
                    audio_dict = {}
                    for _, row in df_audio.dropna(subset=['audio_id', 'item']).iterrows():
                        aid = str(row['audio_id']).strip().lower()
                        item_val = str(row['item']).strip()
                        if aid and aid != 'nan':
                            # 安全地移除常見副檔名，不要用 os.path.splitext 避免 "01.檔名" 被從中間切斷
                            for ext in [".wav", ".mp3", ".m4a", ".flac"]:
                                if aid.endswith(ext):
                                    aid = aid[:-len(ext)]
                                    break
                            # ✅ 同步截斷 Excel audio_id 裡的 _transcript 後綴
                            # 例如 Excel 填的是 "02.ABA..._transcript_openaiwhisper_medium_正常"
                            # 必須跟 json_base 一樣只保留 _transcript 之前的部分
                            if "_transcript" in aid:
                                aid = aid.split("_transcript")[0]
                            audio_dict[aid] = item_val
                    print(f"audio_dict: {audio_dict}")

                    for jf in json_files:
                        jname = jf.name.lower()
                        # 將 JSON 檔名還原成原本的音檔名 (拿掉 _transcript_... 或 .json)
                        json_base = jname.split("_transcript")[0].replace(".json", "")
                        print(f"json_base: {json_base}")
                        
                        # 直接用 json 主檔名去字典裡面查 (如果完全命中就拿，否則預設未知)
                        item = audio_dict.get(json_base, "未知")
                        print(f"item: {item}")
                        
                        # # 如果完全命中失敗，退回「包含」比對 (以防 json 檔名跟 excel 有些微出入)
                        # if item == "未知":
                        #     for aid_excel, item_excel in audio_dict.items():
                        #         # 雙向比對，但加上長度防護，避免 "01" 等過短字串造成誤判
                        #         if len(aid_excel) > 4 and (aid_excel in json_base or json_base in aid_excel):
                        #             item = item_excel
                        #             print(f"item: {item_excel}")
                        #             break
                                
                        file_q = num_fixed + flex_counts.get(item, 0)
                        print(f"num_fixed + flex_counts = {num_fixed} + {flex_counts.get(item, 0)} = {file_q}")
                        file_q_counts[jname] = file_q  # 把這份音檔的題數存起來
                        total_q += file_q
                        
                    # 4. 定義各模型單題預估時間 (秒) - 依實際測試經驗設定
                    model_avg_times = {
                        "Gemini": 6.0,
                        "AzureGPT": 3.0,
                    }
                    
                    choice = st.session_state.get("qa_model_choice", "Gemini")
                    avg_time = model_avg_times.get(choice, 10.0)
                    total_wait_sec = total_q * avg_time
                    
                    # 轉換為分鐘和秒以便閱讀
                    wait_minutes = int(total_wait_sec // 60)
                    wait_seconds = int(total_wait_sec % 60)
                    time_str = f"{wait_minutes} 分 {wait_seconds} 秒" if wait_minutes > 0 else f"{wait_seconds} 秒"

                    # st.info(
                        # f"⏱️ **預估等待時間**：質檢題數 (**{total_q} 題**) × 平均單題質檢時間 (**{avg_time} 秒**) = 預估等待 **{time_str}**"
                    # )
                except Exception as e:
                    # If calculation fails, ensure button is re-enabled
                    if "qa_btn_disabled" not in st.session_state:
                        st.session_state.qa_btn_disabled = False
                    st.session_state.qa_btn_disabled = False
                    st.rerun() # Rerun to re-enable the button if an error occurs here
                    
            if "qa_btn_disabled" not in st.session_state:
                st.session_state.qa_btn_disabled = False
            # 切換 tab 再回來：若上次質檢已完成，自動解鎖讓使用者可以重新執行
            elif st.session_state.get("qa_done", False):
                st.session_state.qa_btn_disabled = False
                
            def disable_qa_btn():
                st.session_state.qa_btn_disabled = True

            # 等待時間動態顯示目前已移到 Spinner 內部
            
            # ========== Step 7: 執行按鈕 ==========
            if st.button("🚀 執行質檢", key="qa_start_button", disabled=st.session_state.qa_btn_disabled, on_click=disable_qa_btn):
                # 檢查必要條件
                if not st.session_state.get("acceptance_id"):
                    st.error("❌ 請先建立專案（Step 0）")
                    st.session_state.qa_btn_disabled = False # Re-enable button on error
                    st.stop()
                elif not st.session_state.get("qa_question_path") or not st.session_state.get("qa_audioitem_path"):
                    st.error("❌ 請先上傳 audio_item 和 questionset")
                    st.session_state.qa_btn_disabled = False # Re-enable button on error
                    st.stop()
                elif not json_files:
                    st.error("❌ 專案資料夾中沒有找到逐字稿 JSON 檔案")
                    st.session_state.qa_btn_disabled = False # Re-enable button on error
                    st.stop()
                else:
                    # ── 每次執行前清空上一次的質檢結果 ──
                    for _k in [
                        "qa_done", "qa_results",
                        "qa_show_filter",
                        "qa_confirm_query", "qa_confirmed_status", "qa_confirmed_audio_ids",
                        "qa_filter_confirmed", "qa_filter_excel_bytes", "qa_filter_export_filename",
                        "qa_filter_export_path", "qa_filter_export_info",
                        "qa_current_page", "qa_last_filter",
                    ]:
                        st.session_state.pop(_k, None)

                    # 初始化模型
                    try:
                        if st.session_state.qa_model_choice == "Gemini":
                            genai.configure(api_key=os.environ.get("GOOGLE_API_KEY", ""))
                            model = genai.GenerativeModel("gemini-2.5-flash")
                            st.info("✅ Gemini 模型已初始化")
                        elif st.session_state.qa_model_choice == "AzureGPT":
                            model = qa_agent.AzureGPTModel()
                            st.info(f"✅ AzureGPT 模型已初始化（{os.getenv('AZURE_DEPLOYMENT', '')}）")


                        # 用於檔名的 LLM 標籤
                        _mc = st.session_state.qa_model_choice
                        if _mc == "Gemini":
                            llm_label = "Gemini"
                        elif _mc == "AzureGPT":
                            llm_label = "AzureGPT"
                        else:
                            llm_label = _mc
                        st.session_state["qa_llm_label"] = llm_label

                        # Terminal 印出 / 操作紀錄
                        print(f"[QA] 使用模型：{llm_label}")
                        
                        # 設定輸出目錄為專案資料夾的 qa_results 子目錄
                        project_folder = init_project_log_folder(st.session_state.acceptance_id)
                        qa_output_dir = os.path.join(project_folder, "qa_results")
                        os.makedirs(qa_output_dir, exist_ok=True)

                        # 讀取剛剛上面計算出來的預估時間參數（若存在）
                        total_q_display = vars().get('total_q', '未知')
                        avg_time_display = vars().get('avg_time', '未知')
                        time_str_display = vars().get('time_str', '未知')
                        
                        spinner_msg = f"⏱️ 正在執行 AI 質檢... 預估等待時間：質檢題數 ({total_q_display} 題) × 平均單題質檢時間 ({avg_time_display} 秒) = 預估等待 {time_str_display}"
                        
                        # 執行批次質檢
                        append_log(st.session_state.acceptance_id, f"開始批次質檢，使用模型：{llm_label}，共 {len(json_files)} 個檔案")
                        
                        with st.spinner(spinner_msg):
                            results = []
                            progress_bar = st.progress(0)
                            
                            import time
                            
                            for idx, json_file in enumerate(json_files, 1):
                                try:
                                    st.caption(f"處理中 [{idx}/{len(json_files)}]: {json_file.name}")
                                    
                                    # 開始計時
                                    start_time = time.time()
                                    
                                    result = qa_agent.run_qa_on_transcript_json(
                                        INPUT_JSON=str(json_file),
                                        model=model,
                                        question_set_path=st.session_state.qa_question_path,
                                        audio_item_path=st.session_state.qa_audioitem_path,
                                        output_dir=qa_output_dir
                                    )
                                    
                                    # 結束計時
                                    end_time = time.time()
                                    elapsed = end_time - start_time
                                    
                                    # 讀取該檔案剛剛被估算的題數
                                    # 注意：這裡使用剛剛在前一個 if 區塊算出的 file_q_counts 字典
                                    q_counts_dict = vars().get('file_q_counts', {})
                                    this_file_q = q_counts_dict.get(json_file.name.lower(), '未知')
                                    
                                    # 印出實際花費時間與總題數
                                    msg = f"[系統實測] 檔案 {json_file.name} 總題數: {this_file_q} 質檢花費時間: {elapsed:.2f} 秒"
                                    print(msg)
                                    append_log(st.session_state.acceptance_id, msg)
                                    
                                    results.append(result)
                                    progress_bar.progress(idx / len(json_files))
                                except Exception as e:
                                    st.warning(f"⚠️ 處理 {json_file.name} 時發生錯誤：{e}")
                                    append_log(st.session_state.acceptance_id, f"質檢錯誤 ({json_file.name})：{e}")
                            
                            st.session_state.qa_results = results
                            st.session_state.qa_done = True
                            append_log(st.session_state.acceptance_id, f"批次質檢完成，成功處理 {len(results)} 個檔案，結果儲存至：{qa_output_dir}")
                            st.success(f"✅ 批次質檢完成！成功處理 {len(results)} 個檔案")#\n\n結果已儲存至：{qa_output_dir}
                            
                            # 自動生成質檢報告 Excel
                            try:
                                st.info("📊 正在生成質檢報告 Excel...")
                                
                                # 讀取 audioitem.xlsx 建立 audio_id 到 item 的對應
                                audioitem_df = pd.read_excel(st.session_state.qa_audioitem_path, engine='openpyxl')
                                audio_item_map = dict(zip(audioitem_df['audio_id'], audioitem_df['item']))
                                
                                # 準備報告資料
                                report_data = []
                                question_number = 1
                                
                                for result in results:
                                    audio_id = result.get("id", "未知")
                                    answers = result.get("acquired_answer", [])
                                    
                                    # 從 audioitem.xlsx 中查找作業項目
                                    item = audio_item_map.get(audio_id, "未知")
                                    
                                    for ans in answers:
                                        # 找到 answer_key 和 answer_value
                                        answer_value = None
                                        for k, v in ans.items():
                                            if k.startswith("answer_"):
                                                answer_value = v
                                                break
                                        
                                        # 取得是否題屬性
                                        is_yesno = ans.get("yesno", False)
                                        
                                        report_data.append({
                                            "音檔名稱": audio_id,
                                            "作業項目": item,
                                            "質檢題組": ans.get("question_category", ""),
                                            "質檢問題": ans.get("question", ""),
                                            "是非題": is_yesno,
                                            "AI回答": answer_value if answer_value else "",
                                            "AI理由": ans.get("reason", ""),
                                            "逐字稿時間": ans.get("evidence", [])  # 稍後在 DataFrame 中 format
                                        })
                                
                                # 建立 DataFrame
                                report_df = pd.DataFrame(report_data)
                                


                                # 先格式化 逐字稿時間
                                report_df['逐字稿時間'] = report_df['逐字稿時間'].apply(fmt_evidence)

                                # ── Step 1：修正閘門題的 AI回答（沒有「投資型」關鍵字 → 強制改「否」）──
                                gate_mask = report_df['質檢問題'].str.contains('是否為投資型保單', na=False)
                                needs_fix = gate_mask & (report_df['AI回答'].astype(str) == '是') & \
                                            (~report_df['逐字稿時間'].str.contains('投資型', na=False))
                                report_df.loc[needs_fix, 'AI回答'] = '否'

                                # ── Step 2：用修正後的 AI回答 建立閘門對照表 ──
                                invest_gate_map = {}
                                if gate_mask.any():
                                    invest_gate_map = (
                                        report_df[gate_mask]
                                        .set_index('音檔名稱')['AI回答']
                                        .astype(str)
                                        .to_dict()
                                    )

                                # ── Step 3：計算檢核點 & AI質檢結果 ──
                                report_df['檢核點'] = report_df.apply(get_checkpoints, axis=1)
                                report_df['AI質檢結果'] = report_df.apply(lambda r: get_qa_result(r, invest_gate_map), axis=1)

                                # 選取並排序欄位
                                report_df = report_df[[
                                    '音檔名稱', '作業項目', '質檢題組', '質檢問題',
                                    '是非題', '檢核點', 'AI回答', 'AI理由', '逐字稿時間', 'AI質檢結果'
                                ]]
                                
                                
                                # 儲存到專案資料夾的 qa_results（加入時間戳記）
                                qa_report_filename = f"質檢結果_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{st.session_state.get('qa_llm_label', 'LLM')}.xlsx"
                                report_path = os.path.join(qa_output_dir, qa_report_filename)
                                report_df.to_excel(report_path, index=False, engine='openpyxl')
                                st.session_state["qa_report_path"] = report_path
                                
                                # ── 把題組路徑寫入 meta.json ──
                                try:
                                    _meta_path = os.path.join(project_folder, "meta.json")
                                    _meta = {}
                                    if os.path.exists(_meta_path):
                                        with open(_meta_path, "r", encoding="utf-8") as _f:
                                            _meta = json.load(_f)
                                    _meta["qa_audioitem_path"]  = str(st.session_state.get("qa_audioitem_path", ""))
                                    _meta["qa_question_path"]   = str(st.session_state.get("qa_question_path", ""))
                                    _meta["qa_report_path"]     = report_path
                                    _meta["qa_llm"]             = llm_label
                                    with open(_meta_path, "w", encoding="utf-8") as _f:
                                        json.dump(_meta, _f, ensure_ascii=False, indent=2)
                                except Exception as _me:
                                    print(f"⚠️ 無法更新 meta.json QA 路徑：{_me}")
                                
                                # 寫入 log
                                append_log(st.session_state.acceptance_id, f"自動生成質檢報告：{report_path}")
                                
                                st.success(f"✅ 質檢報告已自動生成並儲存至：\n\n{report_path}")
                                
                            except Exception as e:
                                st.warning(f"⚠️ 質檢報告生成失敗：{e}")
                                append_log(st.session_state.acceptance_id, f"質檢報告生成錯誤：{e}")
                    
                    except Exception as e:
                        st.error(f"❌ 質檢過程發生錯誤：{e}")
                        append_log(st.session_state.acceptance_id, f"質檢錯誤：{e}")
                        import traceback
                        st.code(traceback.format_exc())
                        
                st.session_state.qa_btn_disabled = False
                st.rerun()
            # ========== Step 8: 結果顯示 ==========
            st.markdown('----------------------------------')
            
            if st.session_state.get("qa_done", False):
                results = st.session_state.get("qa_results", [])
                
                if results:
                    st.success(f"✅ 質檢完成，共處理 {len(results)} 個檔案")
                    
                    # ── 質檢結果路徑 ──
                    report_path = st.session_state.get("qa_report_path", "")
                    
                    st.markdown('<p style="font-size: 24px; font-weight: bold;">質檢結果路徑</p>', unsafe_allow_html=True)
                    st.table(pd.DataFrame([{"質檢報告路徑": report_path}]))
                    
                    # ── 立即下載按鈕 ──
                    if report_path and os.path.exists(report_path):
                        with open(report_path, "rb") as f:
                            st.download_button(
                                label="⬇️⬇️⬇️ 立即下載質檢結果 Excel",
                                data=f.read(),
                                file_name=os.path.basename(report_path),
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                    else:
                        st.info("質檢報告尚未生成，請先執行質檢")
                    
                    # ── 使用篩選器按鈕 ──
                    if st.button("使用篩選器", key="qa_show_filter_btn"):
                        st.session_state["qa_show_filter"] = True
                    
                    # ── 篩選器區域（點擊後才顯示）──
                    if st.session_state.get("qa_show_filter", False):
                        
                        # Step 1：將結果轉換為 DataFrame 並套用格式化邏輯
                        # 這是為了讓 UI 顯示的狀態與 Excel 報告一致（摘要和詳細結果共用）
                        all_items_data = []
                        for result in results:
                            audio_id = result.get("id", "未知")
                            for ans in result.get("acquired_answer", []):
                                answer_value = None
                                for k, v in ans.items():
                                    if k.startswith("answer_"):
                                        answer_value = v
                                        break
                                
                                all_items_data.append({
                                    "audio_id": audio_id,
                                    "ans": ans,
                                    "質檢題組": ans.get("question_category", ""),
                                    "質檢問題": ans.get("question", ""),
                                    "是非題": ans.get("yesno", False),
                                    "AI回答": answer_value if answer_value else "",
                                    "逐字稿時間": ans.get("evidence", [])
                                })
                        
                        df_ui = pd.DataFrame(all_items_data)
                        
                        # 在 DataFrame 建立音檔名稱的 mapping (因為 get_qa_result 需要 '音檔名稱')
                        df_ui['音檔名稱'] = df_ui['audio_id']
                        
                        # 處理閘門題修補與最終狀態
                        if not df_ui.empty:
                            df_ui['_evidence_str'] = df_ui['逐字稿時間'].apply(lambda evs: " ".join(e.get('text', '') for e in evs if isinstance(e, dict)))
                            gate_mask = df_ui['質檢問題'].str.contains('是否為投資型保單', na=False)
                            needs_fix = gate_mask & (df_ui['AI回答'].astype(str) == '是') & \
                                        (~df_ui['_evidence_str'].str.contains('投資型', na=False))
                            df_ui.loc[needs_fix, 'AI回答'] = '否'
                            
                            invest_gate_map = {}
                            if gate_mask.any():
                                invest_gate_map = (
                                    df_ui[gate_mask]
                                    .set_index('音檔名稱')['AI回答']
                                    .astype(str)
                                    .to_dict()
                                )
                            
                            df_ui['status'] = df_ui.apply(lambda r: get_qa_result(r, invest_gate_map), axis=1)
                            df_ui['檢核點'] = df_ui.apply(get_checkpoints, axis=1)

                        # 摘要結果
                        st.markdown('<p style="font-size: 24px; font-weight: bold;">摘要</p>', unsafe_allow_html=True)
                        if not df_ui.empty:
                            # 依據 audio_id 分組統計各種狀態的數量
                            summary_df = df_ui.groupby('audio_id')['status'].value_counts().unstack(fill_value=0).reset_index()
                            summary_df.rename(columns={'audio_id': 'Audio ID'}, inplace=True)
                            
                            # 確保必要的欄位存在
                            for col in ['正常', '異常', '不用分正常異常']:
                                if col not in summary_df.columns:
                                    summary_df[col] = 0
                                    
                            # 計算總題數
                            summary_df['總題數'] = summary_df['正常'] + summary_df['異常'] + summary_df['不用分正常異常']
                            
                            # 重新排列欄位順序以利閱讀
                            summary_df = summary_df[['Audio ID', '總題數', '正常', '異常', '不用分正常異常']]
                            st.dataframe(summary_df, use_container_width=True)
                        else:
                            st.info("沒有分析資料")
                        
                        # 詳細結果（分頁 + 篩選）
                        st.markdown('<p style="font-size: 24px; font-weight: bold;">篩選器</p>', unsafe_allow_html=True)
                        
                        # 轉回 list of dict 給原本的 UI 渲染邏輯用
                        all_items = df_ui.to_dict('records') if not df_ui.empty else []
                        
                        # Step 2：篩選控制（狀態 + 音檔名稱）
                        all_audio_ids = sorted(set(item["audio_id"] for item in all_items))
                        
                        filter_col1, filter_col2, filter_col3 = st.columns([2, 4, 1])
                        with filter_col1:
                            selected_status = st.radio(
                                "狀態",
                                options=["全部", "正常", "異常"],
                                horizontal=True,
                                key="qa_filter_status"
                            )
                        with filter_col2:
                            selected_audio_ids = st.multiselect(
                                "音檔名稱（空白 = 全部）",
                                options=all_audio_ids,
                                default=[],
                                key="qa_filter_audio"
                            )
                        
                        # 確認查詢按鈕
                        if st.button("確認查詢", key="qa_confirm_query_btn"):
                            st.session_state["qa_confirm_query"] = True
                            st.session_state["qa_current_page"] = 1
                            # 儲存當下篩選條件
                            st.session_state["qa_confirmed_status"] = selected_status
                            st.session_state["qa_confirmed_audio_ids"] = selected_audio_ids
                        
                        # 只有按下確認查詢後才顯示結果
                        if st.session_state.get("qa_confirm_query", False):
                            # 使用確認時的篩選條件
                            confirmed_status = st.session_state.get("qa_confirmed_status", "全部")
                            confirmed_audio_ids = st.session_state.get("qa_confirmed_audio_ids", [])
                            
                            # 同時套用兩個篩選條件
                            filtered_items = [
                                item for item in all_items
                                if (confirmed_status == "全部" or item["status"] == confirmed_status)
                                and (not confirmed_audio_ids or item["audio_id"] in confirmed_audio_ids)
                            ]
                            
                            total_items = len(filtered_items)
                            PAGE_SIZE = 10
                            total_pages = max(1, (total_items + PAGE_SIZE - 1) // PAGE_SIZE)
                            
                            with filter_col3:
                                st.markdown(f"共 **{total_items}** 題 / **{total_pages}** 頁")
                            
                            current_page = st.session_state.get("qa_current_page", 1)
                            current_page = max(1, min(current_page, total_pages))
                            
                            # 分頁按鈕列
                            if total_pages > 1:
                                pg_col1, pg_col2, pg_col3 = st.columns([1, 3, 1])
                                with pg_col1:
                                    if st.button("◀ 上一頁", key="qa_prev_page", disabled=(current_page <= 1)):
                                        st.session_state["qa_current_page"] = current_page - 1
                                        st.rerun()
                                with pg_col2:
                                    st.markdown(f"<div style='text-align:center; padding-top:8px;'>第 <b>{current_page}</b> 頁 / 共 <b>{total_pages}</b> 頁</div>", unsafe_allow_html=True)
                                with pg_col3:
                                    if st.button("下一頁 ▶", key="qa_next_page", disabled=(current_page >= total_pages)):
                                        st.session_state["qa_current_page"] = current_page + 1
                                        st.rerun()
                            
                            # 取出當頁資料並顯示
                            start_idx = (current_page - 1) * PAGE_SIZE
                            page_items = filtered_items[start_idx : start_idx + PAGE_SIZE]
                            
                            if not page_items:
                                st.info("沒有符合篩選條件的結果")
                            else:
                                prev_audio_id = None
                                for idx, item in enumerate(page_items):
                                    audio_id = item["audio_id"]
                                    status = item["status"]
                                    status_color = "red" if status == "異常" else "green"
                                    
                                    if audio_id != prev_audio_id:
                                        st.markdown(f"#### 🎵 {audio_id}")
                                        prev_audio_id = audio_id
                                    
                                    q_num = start_idx + idx + 1
                                    st.markdown(f"**Q{q_num}. [{item.get('質檢題組', '')}] {item.get('質檢問題', '')}**")
                                    st.markdown(f":{status_color}[**{status}**]")
                                    st.write(f"回答：{item.get('AI回答', '')}")
                                    st.write(f"理由：{item.get('ans', {}).get('reason', '')}")
                                    
                                    evidence = item.get("逐字稿時間", [])
                                    if evidence:
                                        st.write("逐字稿時間：")
                                        for ev in evidence:
                                            st.caption(f"  - {ev.get('raw_timestamp', '')} {ev.get('text', '')}")
                                    
                                    st.markdown("---")

                            
                            # ── 確認篩選條件 ──
                            if st.button("確認篩選條件", key="qa_filter_confirm_btn"):
                                now_str = datetime.now().strftime("%Y%m%d_%H%M%S")
                                export_filename = f"篩選質檢結果_{now_str}_{st.session_state.get('qa_llm_label', 'LLM')}.xlsx"
                                
                                # 儲存至專案 qa_results 資料夾
                                project_folder = init_project_log_folder(st.session_state.acceptance_id)
                                qa_results_dir = os.path.join(project_folder, "qa_results")
                                os.makedirs(qa_results_dir, exist_ok=True)
                                export_path = os.path.join(qa_results_dir, export_filename)
                                
                                # 將篩選結果轉為 DataFrame（欄位與 qareport.xlsx 統一）
                                export_rows = []
                                for q_num, item in enumerate(filtered_items, 1):
                                    # 使用已格式化過的狀態與檢核點
                                    export_rows.append({
                                        "音檔名稱": item["audio_id"],
                                        "作業項目": item.get("item", ""),  # 若 filter array 沒 item 就為空
                                        "質檢題組": item.get('質檢題組', ''),
                                        "質檢問題": item.get('質檢問題', ''),
                                        "是非題": item.get('是非題', False),
                                        "檢核點": item.get('檢核點', ''),
                                        "AI回答": item.get('AI回答', ''),
                                        "AI理由": item.get('ans', {}).get("reason", ""),
                                        "逐字稿時間": fmt_evidence(item.get("逐字稿時間", [])),
                                        "AI質檢結果": item["status"]
                                    })
                                # EXPORT_COLUMNS 對齊主要報表
                                EXPORT_COLUMNS = ['音檔名稱', '作業項目', '質檢題組', '質檢問題', '是非題', '檢核點', 'AI回答', 'AI理由', '逐字稿時間', 'AI質檢結果']
                                export_df = pd.DataFrame(export_rows, columns=EXPORT_COLUMNS) if export_rows else pd.DataFrame(columns=EXPORT_COLUMNS)

                                
                                # 存到記憶體
                                excel_bytes = io.BytesIO()
                                export_df.to_excel(excel_bytes, index=False)
                                excel_bytes.seek(0)
                                
                                # 存到本機
                                with open(export_path, "wb") as f:
                                    f.write(excel_bytes.getvalue())
                                
                                # 存到 session_state
                                st.session_state["qa_filter_excel_bytes"] = excel_bytes.getvalue()
                                st.session_state["qa_filter_export_filename"] = export_filename
                                st.session_state["qa_filter_export_path"] = export_path
                                st.session_state["qa_filter_export_info"] = pd.DataFrame([{
                                    "篩選質檢結果路徑": export_path,
                                    "篩選時間": datetime.now().strftime("%Y/%m/%d %H:%M:%S")
                                }])
                                st.session_state["qa_filter_confirmed"] = True
                            
                            # ── 篩選完成後永遠顯示路徑 + 下載 ──
                            if st.session_state.get("qa_filter_confirmed", False):
                                st.markdown('<p style="font-size: 24px; font-weight: bold;">篩選質檢結果路徑</p>', unsafe_allow_html=True)
                                st.table(st.session_state.get("qa_filter_export_info", pd.DataFrame()))
                                st.success("篩選完成，已儲存至篩選質檢結果路徑")
                                st.download_button(
                                    label="⬇️⬇️⬇️ 立即下載篩選質檢結果 Excel",
                                    data=st.session_state["qa_filter_excel_bytes"],
                                    file_name=st.session_state.get("qa_filter_export_filename", "篩選質檢結果.xlsx"),
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    key="qa_filter_download_btn"
                                )

            else:
                st.info("尚未執行質檢，請設定模型、題組後，點擊「執行質檢」按鈕")





# ➤ 4. 歷史結果查看頁面
elif st.session_state.project_type == "歷史結果查看":
    st.markdown('<p style="font-size: 36px; font-weight: bold;">📋 歷史結果查看</p>', unsafe_allow_html=True)
    print("✅ 進入歷史結果查看頁面")

    # Step 1：初始化 df_history、history_filters
    def extract_date_unit(acceptance_id):# 新增欄位：解析受理編號產出「建立日期」與「受理單位」
        parts = acceptance_id.split("_")
        date = parts[0]
        unit_code = parts[-1]
        unit_map = {"1": "服務中心", "2": "智能客服中心", "3": "直效行銷部"}
        return date, unit_map.get(unit_code, "未知")
    
    st.session_state.history = load_history_from_logs()
    df_history = pd.DataFrame(st.session_state.history)
    print(df_history.columns)
    # expected_cols = ["受理編號", "員工編號", "音檔路徑", "切角色模式", "逐字稿路徑", "關鍵字路徑", "比對結果", "專案預覽"]


    if not df_history.empty:
        df_history["建立日期"], df_history["受理單位"] = zip(*df_history["受理編號"].map(extract_date_unit))
    else:
        df_history["建立日期"] = []
        df_history["受理單位"] = []

    if "history_filters" not in st.session_state:
        print(st.session_state.get("employee_id", "尚未輸入"))
        st.session_state.history_filters = {
            "date": "",
            "empid": st.session_state.get("employee_id", "all"),
            "unit": "all",
            "id": "all"
        }
    else:
        # 若 filters 是 all 而且 employee_id 有指定，就預設填入
        if (
            st.session_state.history_filters.get("empid") == "all" and
            st.session_state.get("employee_id", "all") != "all"
        ):
            st.session_state.history_filters["empid"] = st.session_state.get("employee_id")

    if "history_query_triggered" not in st.session_state:
        st.session_state.history_query_triggered = False

    # Step 2：建立四個交叉過濾的選項欄位
    def get_filtered_df(df, filters):
        df_filtered = df.copy()
        if filters["date"]:
            df_filtered = df_filtered[df_filtered["建立日期"].str.startswith(filters["date"])]
        if filters["empid"] != "all":
            df_filtered = df_filtered[df_filtered["員工編號"] == filters["empid"]]
        if filters["unit"] != "all":
            df_filtered = df_filtered[df_filtered["受理單位"] == filters["unit"]]
        if filters["id"] != "all":
            df_filtered = df_filtered[df_filtered["受理編號"] == filters["id"]]
        return df_filtered

    filters = st.session_state.history_filters
    col1, col2, col3, col4 = st.columns(4)

    # 建立日期
    with col1:
        date_input = st.text_input("建立日期（例如 202507）", value=filters["date"])

    # 員工編號（根據當前「建立日期」過濾）
    with col2:
        df_empid = get_filtered_df(df_history, {
            "date": date_input, "empid": "all", "unit": "all", "id": "all"
        }) #用日期產出小df
        empid_options = ["all"] + sorted(df_empid["員工編號"].dropna().unique()) if "員工編號" in df_empid.columns else ["all"]

        # 嘗試從 session 中取得 employee_id
        session_empid = st.session_state.get("employee_id", "all")
        print(f"session_empid:{session_empid}")

        if session_empid not in empid_options:
            session_empid = "all"

        empid_input = st.selectbox("員工編號", empid_options, #小df員工編號候選清單
                                index=empid_options.index(session_empid)) 
        print(f"empid_input:{empid_input}")
        print(f"empid_option:{empid_options}")
        print(f"empid_input_index:{empid_options.index(session_empid)}")

    # 受理單位（根據當前「建立日期」＋「員編」過濾）
    with col3:
        df_unit = get_filtered_df(df_history, {
            "date": date_input, "empid": empid_input, "unit": "all", "id": "all"
        })
        unit_options = ["all"] + sorted(df_unit["受理單位"].dropna().unique()) if "受理單位" in df_unit.columns else ["all"]
        unit_input = st.selectbox("受理單位", unit_options,
                                index=unit_options.index(filters["unit"]) if filters["unit"] in unit_options else 0) #如果 filters["unit"] 有在選單中，選擇該項；否則預設選 0（也就是 "all"）

    # 受理編號（根據前三欄過濾）
    with col4:
        df_id = get_filtered_df(df_history, {
            "date": date_input, "empid": empid_input, "unit": unit_input, "id": "all"
        })
        id_options = ["all"] + sorted(df_id["受理編號"].dropna().unique()) if "受理編號" in df_id.columns else ["all"]
        id_input = st.selectbox("受理編號", id_options,
                                index=id_options.index(filters["id"]) if filters["id"] in id_options else 0)
    
    # Step 5：如果任一欄位改變，更新 session 並 rerun
    new_filters = {
        "date": date_input,
        "empid": empid_input,
        "unit": unit_input,
        "id": id_input        
    }

    if new_filters != st.session_state.history_filters:
        st.session_state.history_filters = new_filters
        st.session_state.history_query_triggered = False

    # Step 5：確認查詢按鈕
    if st.button("確認查詢"):
        print("✅ 查詢結果已觸發，準備畫表格和按鈕")
        st.session_state.history_query_triggered = True

    # Step 6：顯示查詢結果
    if st.session_state.history_query_triggered:
        final_filtered_df = get_filtered_df(df_history, st.session_state.history_filters)
        if not final_filtered_df.empty:
            for col in ["關鍵字路徑", "比對結果路徑"]:
                final_filtered_df[col] = final_filtered_df[col].apply(local_path_to_download_button)
            final_filtered_df["專案預覽"] = final_filtered_df["專案預覽"].apply(local_path_to_http)
            # 題組：多個路徑組合成小 HTML
            def fmt_topic_paths(paths):
                if not paths:
                    return "-"
                return "<br>".join(local_path_to_download_button(p) for p in paths if p)
            final_filtered_df["題組"] = final_filtered_df["題組路徑清單"].apply(fmt_topic_paths)
            # 質檢結果：單個路徑
            final_filtered_df["質檢結果"] = final_filtered_df["質檢結果路徑"].apply(
                lambda p: local_path_to_download_button(p) if p else "-"
            )
        desired_columns = ["建立日期", "員工編號", "受理單位", "專案預覽", "關鍵字路徑", "比對結果路徑", "題組", "質檢結果"]
        # 表頭別名（只影響顯示）
        header_alias = {
            "關鍵字路徑": "關鍵字",
            "比對結果路徑": "比對結果",
            "題組": "題組",
            "質檢結果": "質檢結果",
        }        
        if not final_filtered_df.empty and all(col in final_filtered_df.columns for col in desired_columns):
            # 顯示表格標題欄
            header_cols = st.columns(len(desired_columns) + 1)  # 多一欄放「編輯」按鈕
            for i, col_name in enumerate(desired_columns):
                header_cols[i].markdown(f"**{header_alias.get(col_name, col_name)}**")#(f"**{col_name}**")
            header_cols[-1].markdown("**操作**")  # 顯示編輯欄位標題

            # 逐列顯示資料 + 編輯按鈕
            for i, row in final_filtered_df.iterrows():
                cols = st.columns(len(desired_columns) + 1)  # 同樣多一欄放「編輯」按鈕

                for j, col_name in enumerate(desired_columns):
                    if col_name in ["專案預覽", "關鍵字路徑", "比對結果路徑", "題組", "質檢結果"]:
                        cols[j].markdown(row[col_name], unsafe_allow_html=True)  # HTML欄位用markdown顯示
                    else:
                        cols[j].write(str(row[col_name]))  # 其他欄位正常顯示
                
                # 建立唯一 key 避免重複錯誤
                edit_key = f"edit_{row['受理編號']}"

                if cols[-1].button("編輯", key=edit_key):  # 點編輯按鈕
                    print('✅ 點編輯按鈕')
                    # 1) 宣告 clone 模式
                    st.session_state.reuse_mode = "audio_and_transcript"
                    st.session_state["is_clone"] = True                  

                    # 2) 準備這次要 clone 的資料
                    st.session_state.reuse_project_data = {
                        "原始受理編號": row.get("受理編號",""),
                        "員工編號": row.get("員工編號",""),
                        "受理單位": row.get("受理單位","未知"),
                        "受理單位代碼": row.get("受理單位代碼","4"),
                        "切角色模式": row.get("切角色模式"),
                        "音檔路徑清單": row.get("音檔路徑清單"),
                        "逐字稿路徑清單": row.get("逐字稿路徑清單"),
                        "關鍵字路徑": row.get("關鍵字路徑"),
                        "比對結果路徑": row.get("比對結果路徑"),
                    }              
                    print(f"原始受理編號:{st.session_state.reuse_project_data['原始受理編號']}")
                    print(f"reuse_data:{st.session_state.reuse_project_data}")

                    # 3) 這是一個全新的 clone 流程 → 重置流程控制與殘留
                    st.session_state.step = 0
                    st.session_state.stt_done = False
                    for k in [
                        "acceptance_id","audio_folder",
                        "audio_file_paths","transcript_paths",
                        "uploaded_keywords_path","result_path",
                        "df_results","show_results","log_header_written"
                    ]:
                        st.session_state.pop(k, None)

                    # 儲存重用資料到 session_state
                    st.session_state.clone_initialized = False
                    st.session_state.clone_kw_initialized = False
                    # 3) 讓下一輪「允許預填一次」+ 標記來源
                    st.session_state.clone_prefilled = False
                    st.session_state.clone_source = st.session_state.reuse_project_data['原始受理編號']


                    # 設定跳轉參數
                    st.session_state.force_page_jump = True  # 加入跳轉 flag
                    st.query_params["page"] = "開新專案"
                    st.rerun()

        else:
            st.warning("⚠️ 找不到符合條件的歷史檔案，請確認查詢條件。")